from django.contrib import admin
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.admin import SimpleListFilter
from django.utils.html import format_html
from django.utils.formats import number_format
from django import forms
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
from django.contrib import messages
from django.db.models import Count, Sum, Q
from django.utils import timezone
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .models import (
    User, Company, Contact, Note, Task, AuditLog,
    Opportunity, OpportunityActivity, Contract, Invoice, Zone, ContractZone,
    EmailTemplate, EmailLog, EmailCampaign, CampaignRecipient, DocumentAttachment,
    Quote, QuoteLineItem, QuoteAttachment, QuoteActivity,
    EmailSequence, SequenceStep, SequenceEnrollment, SequenceStepExecution,
    CustomerSegment, Ticket, TicketComment, TicketAttachment,
    KBCategory, KBTag, KBArticle, KBArticleView, KBArticleRating,
    KBArticleRelation, KBArticleAttachment, TicketKBArticle,
    Device, StaticDocument,
    CorporatePdfTemplate, ContractTemplate, ServicePackageItem, ContractDocument,
    SeasonalTriggerDate, ZoneOfflineAlert
)


@admin.register(User)
class UserAdmin(BaseUserAdmin):
    fieldsets = BaseUserAdmin.fieldsets + (
        ('CRM Info', {'fields': ('role', 'phone', 'department', 'last_login_ip')}),
        ('Email SMTP Configuration', {
            'fields': ('smtp_email', 'smtp_password'),
            'description': 'Configure Gmail credentials for sending emails as this user. Get app password from: https://myaccount.google.com/apppasswords'
        }),
    )
    list_display = ['username', 'email', 'role', 'is_active', 'date_joined']
    list_filter = ['role', 'is_active', 'date_joined']
    search_fields = ['username', 'email', 'first_name', 'last_name']

    def get_form(self, request, obj=None, **kwargs):
        """Customize form to use password widget for smtp_password field"""
        form = super().get_form(request, obj, **kwargs)
        if 'smtp_password' in form.base_fields:
            # Use PasswordInput with render_value=True to show asterisks but allow viewing
            form.base_fields['smtp_password'].widget = forms.PasswordInput(
                attrs={'autocomplete': 'new-password'},
                render_value=True
            )
        return form


class ContactInline(admin.TabularInline):
    model = Contact
    extra = 0
    fields = ['name', 'email', 'title', 'is_active']

    def get_queryset(self, request):
        """Optimize inline queries"""
        return super().get_queryset(request).select_related('company')


class NoteInline(admin.TabularInline):
    model = Note
    extra = 0
    fields = ['title', 'text', 'note_type', 'priority']

    def get_queryset(self, request):
        """Optimize inline queries"""
        return super().get_queryset(request).select_related('company', 'author', 'contact')


class TaskInline(admin.TabularInline):
    model = Task
    extra = 0
    fields = ['title', 'assigned_to', 'priority', 'status', 'due_date']

    def get_queryset(self, request):
        """Optimize inline queries"""
        return super().get_queryset(request).select_related('company', 'assigned_to', 'created_by')


# SubscriptionPlanInline removed - use Contract model with service_type instead


class CompanyZoneInline(admin.TabularInline):
    model = Zone
    extra = 0
    fields = ['name', 'currently_playing_display', 'status_badge_inline', 'device_name', 'last_seen_online']
    readonly_fields = ['name', 'currently_playing_display', 'status_badge_inline', 'last_seen_online', 'platform']
    verbose_name = "Music Zone"
    verbose_name_plural = "Music Zones (auto-synced from Soundtrack)"
    can_delete = False
    
    def currently_playing_display(self, obj):
        """Display what's currently playing instead of zone ID"""
        if obj.api_raw_data and obj.api_raw_data.get('currently_playing'):
            return obj.api_raw_data.get('currently_playing')
        return "No active playlist/schedule"
    currently_playing_display.short_description = "Currently Playing"
    
    def status_badge_inline(self, obj):
        """Display status with color coding"""
        colors = {
            "online": "green",
            "offline": "red", 
            "no_device": "#DAA520",  # Golden yellow
            "expired": "gray",
            "pending": "gray"
        }
        color = colors.get(obj.status, "gray")
        return format_html(
            "<span style=\"padding: 3px 8px; border-radius: 3px; color: white; background-color: {}; font-size: 11px;\">{}</span>",
            color, obj.get_status_display()
        )
    status_badge_inline.short_description = "Status"
    
    def has_add_permission(self, request, obj):
        # Don't allow manual adding for Soundtrack companies
        return False
    
    def get_queryset(self, request):
        # Only show Soundtrack zones
        qs = super().get_queryset(request)
        return qs.filter(platform='soundtrack')


class BeatBreezeZoneInline(admin.TabularInline):
    model = Zone
    extra = 1
    fields = ['name', 'status', 'device_name', 'notes']
    verbose_name = "Beat Breeze Zone"
    verbose_name_plural = "Beat Breeze Zones (manual entry)"
    
    def get_queryset(self, request):
        # Only show Beat Breeze zones
        qs = super().get_queryset(request)
        return qs.filter(platform='beatbreeze')
    
    def formfield_for_choice_field(self, db_field, request, **kwargs):
        if db_field.name == 'platform':
            kwargs['initial'] = 'beatbreeze'
        return super().formfield_for_choice_field(db_field, request, **kwargs)


@admin.register(Company)
class CompanyAdmin(admin.ModelAdmin):
    list_display = ['name', 'country', 'industry', 'location_count', 'music_zone_count', 'soundtrack_status', 'billing_entity', 'corporate_parent_badge', 'parent_company_link', 'child_count']
    list_filter = ['country', 'industry', 'is_active', 'billing_entity', 'is_corporate_parent', 'parent_company', 'created_at']
    list_select_related = ['parent_company']
    search_fields = ['name', 'legal_entity_name', 'website', 'notes', 'it_notes', 'soundtrack_account_id']
    readonly_fields = ['created_at', 'updated_at', 'current_subscription_summary', 'zones_status_summary', 'child_count']
    autocomplete_fields = ['parent_company']
    actions = ['sync_soundtrack_zones', 'preview_email_to_contacts', 'send_email_to_contacts', 'bulk_send_email_to_contacts', 'bulk_sync_soundtrack_zones', 'export_companies_csv', 'export_companies_excel']
    
    def get_inlines(self, request, obj):
        """Show different inlines based on whether company has Soundtrack account"""
        if obj and obj.soundtrack_account_id:
            return [CompanyZoneInline, ContactInline, TaskInline]
        else:
            return [BeatBreezeZoneInline, ContactInline, TaskInline]
    
    def soundtrack_status(self, obj):
        if obj.soundtrack_account_id:
            return format_html('<span style="color: green;">✓ Connected</span>')
        return format_html('<span style="color: gray;">Not connected</span>')
    soundtrack_status.short_description = "Soundtrack"

    def corporate_parent_badge(self, obj):
        """Display badge if company is a corporate parent"""
        if obj.is_corporate_parent:
            return format_html(
                '<span style="background-color: #FFA500; color: white; padding: 3px 8px; '
                'border-radius: 3px; font-size: 11px; font-weight: bold;">CORPORATE HQ</span>'
            )
        return ''
    corporate_parent_badge.short_description = "Corporate Status"

    def parent_company_link(self, obj):
        """Display link to parent company if exists"""
        if obj.parent_company:
            url = reverse('admin:crm_app_company_change', args=[obj.parent_company.pk])
            return format_html('<a href="{}">{}</a>', url, obj.parent_company.name)
        return format_html('<span style="color: #999;">-</span>')
    parent_company_link.short_description = "Parent Company"

    def child_count(self, obj):
        """Display count of child companies for corporate parents"""
        if obj.is_corporate_parent:
            count = obj.child_companies.count()
            if count > 0:
                return format_html(
                    '<span style="background-color: #4CAF50; color: white; padding: 2px 6px; '
                    'border-radius: 3px; font-size: 11px; font-weight: bold;">{} venues</span>',
                    count
                )
            return format_html('<span style="color: #999;">0 venues</span>')
        return format_html('<span style="color: #999;">-</span>')
    child_count.short_description = "Child Companies"

    def zones_status_summary(self, obj):
        """Display summary of zone statuses from Soundtrack API"""
        if not obj.soundtrack_account_id:
            return "No Soundtrack account ID configured"
        
        try:
            from .services.soundtrack_api import soundtrack_api
            # Get zones specifically for this account ID
            api_zones = soundtrack_api.get_account_zones(obj.soundtrack_account_id)
            
            if not api_zones:
                return "No zones found for this account ID"
            
            status_counts = {}
            zone_details = []
            
            for zone in api_zones:
                status = zone.get('status', 'unknown')
                status_counts[status] = status_counts.get(status, 0) + 1
                
                # Create detailed zone info with playlist/schedule
                currently_playing = zone.get('currently_playing', 'No active playlist/schedule')
                zone_details.append(f"• <strong>{zone.get('zone_name', 'Unknown')}</strong>: {currently_playing} <em>({status})</em>")
            
            # Create summary with counts
            summary_parts = []
            for status, count in status_counts.items():
                summary_parts.append(f"{count} {status.title()}")
            
            summary = ", ".join(summary_parts)
            
            # Add detailed zone list
            details = "<br>".join(zone_details)
            
            return format_html(f"{summary}<br><br><small>{details}</small>")
            
        except Exception as e:
            return f"Error fetching zones: {str(e)}"
    zones_status_summary.short_description = "Zone Status Summary"
    
    def sync_soundtrack_zones(self, request, queryset):
        """Sync zones with Soundtrack API for selected companies"""
        from .services.soundtrack_api import soundtrack_api
        
        total_synced = 0
        total_errors = 0
        
        for company in queryset:
            if company.soundtrack_account_id:
                try:
                    synced, errors = soundtrack_api.sync_company_zones(company)
                    total_synced += synced
                    total_errors += errors
                    if synced > 0:
                        self.message_user(request, f"✓ {company.name}: Synced {synced} zones")
                except Exception as e:
                    self.message_user(request, f"✗ {company.name}: Error - {str(e)}", level='ERROR')
                    total_errors += 1
            else:
                self.message_user(request, f"✗ {company.name}: No Soundtrack account ID", level='WARNING')
        
        if total_synced > 0:
            self.message_user(request, f"Successfully synced {total_synced} zones total")
    sync_soundtrack_zones.short_description = "Sync Soundtrack zones"
    
    def preview_email_to_contacts(self, request, queryset):
        """Preview email to contacts of selected companies"""
        if queryset.count() != 1:
            self.message_user(request, 'Please select exactly one company', level='ERROR')
            return

        company = queryset.first()
        # Redirect to preview email form
        from django.urls import reverse
        from django.http import HttpResponseRedirect
        return HttpResponseRedirect(
            reverse('admin_preview_email_company', args=[company.pk])
        )
    preview_email_to_contacts.short_description = 'Preview email to company contacts'

    def send_email_to_contacts(self, request, queryset):
        """Send email to contacts of selected companies"""
        if queryset.count() != 1:
            self.message_user(request, 'Please select exactly one company', level='ERROR')
            return

        company = queryset.first()
        # Redirect to send email form
        from django.urls import reverse
        from django.http import HttpResponseRedirect
        return HttpResponseRedirect(
            reverse('admin_send_email_company', args=[company.pk])
        )
    send_email_to_contacts.short_description = 'Send email to company contacts'
    
    def save_model(self, request, obj, form, change):
        """Auto-sync zones when Soundtrack account ID is added or changed"""
        import logging
        logger = logging.getLogger(__name__)
        
        old_account_id = None
        if change and obj.pk:
            old_obj = Company.objects.get(pk=obj.pk)
            old_account_id = old_obj.soundtrack_account_id
        
        super().save_model(request, obj, form, change)
        
        logger.info(f"Saving company {obj.name}, soundtrack_account_id: {obj.soundtrack_account_id}, old: {old_account_id}")
        
        # If Soundtrack account ID was added or changed, sync zones
        if obj.soundtrack_account_id and obj.soundtrack_account_id != old_account_id:
            from .services.soundtrack_api import soundtrack_api
            try:
                logger.info(f"Attempting to sync zones for {obj.name}")
                synced, errors = soundtrack_api.sync_company_zones(obj)
                logger.info(f"Sync result: {synced} synced, {errors} errors")
                if synced > 0:
                    self.message_user(request, f"✓ Automatically synced {synced} zones from Soundtrack", 'SUCCESS')
                else:
                    self.message_user(request, f"No zones synced. Check the API connection.", 'WARNING')
            except Exception as e:
                logger.error(f"Error syncing zones: {str(e)}")
                self.message_user(request, f"Could not sync zones: {str(e)}", 'WARNING')

    def bulk_send_email_to_contacts(self, request, queryset):
        """Send email to contacts of multiple selected companies"""
        if queryset.count() > 20:
            self.message_user(request, 'Please select 20 companies or fewer for bulk email sending', level='ERROR')
            return

        total_contacts = 0
        companies_processed = 0

        for company in queryset:
            contacts = company.contacts.filter(is_active=True, receives_notifications=True, unsubscribed=False)
            contact_count = contacts.count()

            if contact_count > 0:
                total_contacts += contact_count
                companies_processed += 1
                self.message_user(request, f"✓ {company.name}: {contact_count} contacts ready for email")
            else:
                self.message_user(request, f"✗ {company.name}: No active contacts found", level='WARNING')

        if companies_processed > 0:
            # Redirect to bulk email form
            company_ids = ','.join(str(c.pk) for c in queryset)
            return HttpResponseRedirect(
                reverse('admin_send_email') + f'?company_ids={company_ids}'
            )
        else:
            self.message_user(request, 'No companies with active contacts found', level='ERROR')
    bulk_send_email_to_contacts.short_description = 'Send bulk email to company contacts'

    def bulk_sync_soundtrack_zones(self, request, queryset):
        """Bulk sync zones for multiple companies with Soundtrack accounts"""
        from .services.soundtrack_api import soundtrack_api

        soundtrack_companies = queryset.filter(soundtrack_account_id__isnull=False).exclude(soundtrack_account_id='')

        if not soundtrack_companies.exists():
            self.message_user(request, 'No companies with Soundtrack account IDs found', level='WARNING')
            return

        total_synced = 0
        total_errors = 0
        companies_processed = 0

        for company in soundtrack_companies:
            try:
                synced, errors = soundtrack_api.sync_company_zones(company)
                total_synced += synced
                total_errors += errors
                companies_processed += 1

                if synced > 0:
                    self.message_user(request, f"✓ {company.name}: Synced {synced} zones")
                elif errors > 0:
                    self.message_user(request, f"⚠ {company.name}: {errors} errors during sync", level='WARNING')
                else:
                    self.message_user(request, f"→ {company.name}: No zones to sync")
            except Exception as e:
                self.message_user(request, f"✗ {company.name}: Error - {str(e)}", level='ERROR')
                total_errors += 1

        self.message_user(request, f"Bulk sync completed: {total_synced} zones synced across {companies_processed} companies")
    bulk_sync_soundtrack_zones.short_description = 'Bulk sync Soundtrack zones'

    def export_companies_csv(self, request, queryset):
        """Export selected companies to CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="companies_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'

        writer = csv.writer(response)
        # Write headers
        writer.writerow([
            'Company Name', 'Legal Entity Name', 'Country', 'Industry', 'Website', 'Location Count',
            'Music Zone Count', 'Billing Entity', 'Soundtrack Account ID', 'Contact Count',
            'Active Contract Count', 'Total Contract Value', 'Is Active',
            'Address', 'Created Date', 'Last Updated'
        ])

        # Write data with optimized queries
        companies = queryset.select_related().prefetch_related(
            'contacts', 'contracts'
        ).annotate(
            contact_count=Count('contacts', filter=Q(contacts__is_active=True)),
            active_contract_count=Count('contracts', filter=Q(contracts__is_active=True)),
            total_contract_value=Sum('contracts__value', filter=Q(contracts__is_active=True))
        )

        for company in companies:
            writer.writerow([
                company.name,
                company.legal_entity_name,
                company.country,
                company.get_industry_display() if company.industry else '',
                company.website,
                company.location_count,
                company.music_zone_count,
                company.get_billing_entity_display(),
                company.soundtrack_account_id,
                company.contact_count or 0,
                company.active_contract_count or 0,
                f"{company.total_contract_value or 0:.2f}",
                'Yes' if company.is_active else 'No',
                company.full_address,
                company.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                company.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])

        self.message_user(request, f'Successfully exported {queryset.count()} companies to CSV')
        return response
    export_companies_csv.short_description = 'Export selected companies to CSV'

    def export_companies_excel(self, request, queryset):
        """Export selected companies to Excel with formatting"""
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Companies Export"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        headers = [
            'Company Name', 'Legal Entity Name', 'Country', 'Industry', 'Website', 'Location Count',
            'Music Zone Count', 'Billing Entity', 'Soundtrack Account ID', 'Contact Count',
            'Active Contract Count', 'Total Contract Value (USD)', 'Is Active',
            'Full Address', 'Created Date', 'Last Updated'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data with optimized queries
        companies = queryset.select_related().prefetch_related(
            'contacts', 'contracts'
        ).annotate(
            contact_count=Count('contacts', filter=Q(contacts__is_active=True)),
            active_contract_count=Count('contracts', filter=Q(contracts__is_active=True)),
            total_contract_value=Sum('contracts__value', filter=Q(contracts__is_active=True))
        )

        for row, company in enumerate(companies, 2):
            ws.cell(row=row, column=1, value=company.name)
            ws.cell(row=row, column=2, value=company.legal_entity_name)
            ws.cell(row=row, column=3, value=company.country)
            ws.cell(row=row, column=4, value=company.get_industry_display() if company.industry else '')
            ws.cell(row=row, column=5, value=company.website)
            ws.cell(row=row, column=6, value=company.location_count)
            ws.cell(row=row, column=7, value=company.music_zone_count)
            ws.cell(row=row, column=8, value=company.get_billing_entity_display())
            ws.cell(row=row, column=9, value=company.soundtrack_account_id)
            ws.cell(row=row, column=10, value=company.contact_count or 0)
            ws.cell(row=row, column=11, value=company.active_contract_count or 0)
            ws.cell(row=row, column=12, value=float(company.total_contract_value or 0))
            ws.cell(row=row, column=13, value='Yes' if company.is_active else 'No')
            ws.cell(row=row, column=14, value=company.full_address)
            ws.cell(row=row, column=15, value=company.created_at)
            ws.cell(row=row, column=16, value=company.updated_at)

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = max(len(str(ws.cell(row=1, column=col).value)), 15)
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="companies_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        wb.save(response)
        self.message_user(request, f'Successfully exported {queryset.count()} companies to Excel')
        return response
    export_companies_excel.short_description = 'Export selected companies to Excel'
    
    def current_subscription_summary(self, obj):
        """Display a summary of current subscription plans"""
        plans = obj.subscription_plans.filter(is_active=True)
        if not plans:
            return "No active subscription plans"
        
        summary = []
        for plan in plans:
            # Format dates if they exist
            date_info = ""
            if plan.start_date and plan.end_date:
                date_info = f" ({plan.start_date.strftime('%d/%m/%Y')} - {plan.end_date.strftime('%d/%m/%Y')})"
            elif plan.start_date:
                date_info = f" (from {plan.start_date.strftime('%d/%m/%Y')})"
            
            summary.append(f"{plan.tier}: {plan.zone_count} zones{date_info}")
        
        return " | ".join(summary)
    
    current_subscription_summary.short_description = "Current Subscription Plans"

    def get_queryset(self, request):
        """Optimize queries for list display"""
        return super().get_queryset(request).prefetch_related(
            'contacts', 'zones', 'contracts'
        )
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('name', 'legal_entity_name', 'country', 'industry', 'website', 'location_count', 'music_zone_count'),
        }),
        ('Corporate Structure', {
            'fields': ('is_corporate_parent', 'parent_company', 'child_count'),
            'description': 'Configure corporate parent-child relationships for multi-venue organizations'
        }),
        ('Billing', {
            'fields': ('billing_entity',),
            'description': 'Select which BMAsia legal entity will be used for billing and invoicing'
        }),
        ('Soundtrack Integration', {
            'fields': ('soundtrack_account_id', 'zones_status_summary'),
        }),
        ('Subscription Details', {
            'fields': ('current_subscription_summary',),
        }),
        ('Address', {
            'fields': ('address_line1', 'address_line2', 'city', 'state', 'postal_code'),
            'classes': ('collapse',)
        }),
        ('Notes', {
            'fields': ('notes',),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )


class ContactAdminForm(forms.ModelForm):
    """Custom form for Contact admin with multi-select for notification types"""
    NOTIFICATION_TYPE_CHOICES = [
        ('renewal', 'Renewal Reminders'),
        ('payment', 'Payment Reminders'),
        ('invoice', 'New Invoices'),
        ('quarterly', 'Quarterly Check-ins'),
        ('seasonal', 'Seasonal Campaigns'),
        ('support', 'Support Updates'),
        ('zone_alerts', 'Zone Alerts'),
        ('promotions', 'Promotional Offers'),
        ('newsletters', 'Company Newsletter'),
    ]
    
    notification_types = forms.MultipleChoiceField(
        choices=NOTIFICATION_TYPE_CHOICES,
        widget=forms.CheckboxSelectMultiple,
        required=False,
        help_text="Select which types of automated emails this contact should receive"
    )
    
    class Meta:
        model = Contact
        fields = '__all__'
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.instance and self.instance.pk:
            # Load existing notification types
            self.initial['notification_types'] = self.instance.notification_types or []
    
    def clean_notification_types(self):
        # Convert to list for JSON field
        return self.cleaned_data.get('notification_types', [])


@admin.register(Contact)
class ContactAdmin(admin.ModelAdmin):
    form = ContactAdminForm
    list_display = ['name', 'company', 'email', 'phone', 'contact_type', 'is_primary', 'is_active', 'receives_notifications']
    list_filter = ['contact_type', 'is_primary', 'is_active', 'receives_notifications', 'preferred_language', 'company__country']
    list_select_related = ['company']
    search_fields = ['name', 'email', 'company__name']
    readonly_fields = ['created_at', 'updated_at', 'unsubscribe_token']
    actions = ['bulk_send_email_to_contacts', 'export_contacts_csv', 'export_contacts_excel']
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('company', 'name', 'email', 'phone', 'title', 'department')
        }),
        ('Contact Details', {
            'fields': ('contact_type', 'is_primary', 'is_active', 'linkedin_url')
        }),
        ('Email Preferences', {
            'fields': ('receives_notifications', 'notification_types', 'preferred_language', 'unsubscribed', 'unsubscribe_token'),
            'description': 'Control which automated emails this contact receives'
        }),
        ('Notes', {
            'fields': ('notes', 'last_contacted'),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def bulk_send_email_to_contacts(self, request, queryset):
        """Send email to selected contacts"""
        active_contacts = queryset.filter(is_active=True, receives_notifications=True, unsubscribed=False)

        if not active_contacts.exists():
            self.message_user(request, 'No active contacts found (contacts must be active, receive notifications, and not unsubscribed)', level='WARNING')
            return

        if active_contacts.count() > 50:
            self.message_user(request, 'Please select 50 contacts or fewer for bulk email sending', level='ERROR')
            return

        # Redirect to email form with contact IDs
        contact_ids = ','.join(str(c.pk) for c in active_contacts)
        return HttpResponseRedirect(
            reverse('admin_send_email') + f'?contact_ids={contact_ids}'
        )
    bulk_send_email_to_contacts.short_description = 'Send bulk email to selected contacts'

    def export_contacts_csv(self, request, queryset):
        """Export selected contacts to CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="contacts_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'

        writer = csv.writer(response)
        # Write headers
        writer.writerow([
            'Name', 'Company', 'Email', 'Phone', 'Title', 'Department',
            'Contact Type', 'Is Primary', 'Is Active', 'Receives Notifications',
            'Notification Types', 'Preferred Language', 'LinkedIn URL',
            'Last Contacted', 'Notes', 'Created Date', 'Last Updated'
        ])

        # Write data with optimized queries
        contacts = queryset.select_related('company')

        for contact in contacts:
            writer.writerow([
                contact.name,
                contact.company.name,
                contact.email,
                contact.phone,
                contact.title,
                contact.department,
                contact.get_contact_type_display(),
                'Yes' if contact.is_primary else 'No',
                'Yes' if contact.is_active else 'No',
                'Yes' if contact.receives_notifications else 'No',
                ', '.join(contact.notification_types) if contact.notification_types else '',
                contact.get_preferred_language_display(),
                contact.linkedin_url,
                contact.last_contacted.strftime('%Y-%m-%d %H:%M:%S') if contact.last_contacted else '',
                contact.notes,
                contact.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                contact.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])

        self.message_user(request, f'Successfully exported {queryset.count()} contacts to CSV')
        return response
    export_contacts_csv.short_description = 'Export selected contacts to CSV'

    def export_contacts_excel(self, request, queryset):
        """Export selected contacts to Excel with formatting"""
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contacts Export"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        headers = [
            'Name', 'Company', 'Email', 'Phone', 'Title', 'Department',
            'Contact Type', 'Is Primary', 'Is Active', 'Receives Notifications',
            'Notification Types', 'Preferred Language', 'LinkedIn URL',
            'Last Contacted', 'Notes', 'Created Date', 'Last Updated'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data with optimized queries
        contacts = queryset.select_related('company')

        for row, contact in enumerate(contacts, 2):
            ws.cell(row=row, column=1, value=contact.name)
            ws.cell(row=row, column=2, value=contact.company.name)
            ws.cell(row=row, column=3, value=contact.email)
            ws.cell(row=row, column=4, value=contact.phone)
            ws.cell(row=row, column=5, value=contact.title)
            ws.cell(row=row, column=6, value=contact.department)
            ws.cell(row=row, column=7, value=contact.get_contact_type_display())
            ws.cell(row=row, column=8, value='Yes' if contact.is_primary else 'No')
            ws.cell(row=row, column=9, value='Yes' if contact.is_active else 'No')
            ws.cell(row=row, column=10, value='Yes' if contact.receives_notifications else 'No')
            ws.cell(row=row, column=11, value=', '.join(contact.notification_types) if contact.notification_types else '')
            ws.cell(row=row, column=12, value=contact.get_preferred_language_display())
            ws.cell(row=row, column=13, value=contact.linkedin_url)
            ws.cell(row=row, column=14, value=contact.last_contacted)
            ws.cell(row=row, column=15, value=contact.notes)
            ws.cell(row=row, column=16, value=contact.created_at)
            ws.cell(row=row, column=17, value=contact.updated_at)

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = max(len(str(ws.cell(row=1, column=col).value)), 15)
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="contacts_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        wb.save(response)
        self.message_user(request, f'Successfully exported {queryset.count()} contacts to Excel')
        return response
    export_contacts_excel.short_description = 'Export selected contacts to Excel'

    # list_select_related handles the optimization


@admin.register(Note)
class NoteAdmin(admin.ModelAdmin):
    list_display = ['title', 'company', 'author', 'note_type', 'priority', 'created_at']
    list_filter = ['note_type', 'priority', 'is_private', 'created_at']
    list_select_related = ['company', 'author', 'contact']
    search_fields = ['title', 'text', 'company__name']
    readonly_fields = ['created_at', 'updated_at']


@admin.register(Task)
class TaskAdmin(admin.ModelAdmin):
    list_display = ['title', 'company', 'assigned_to', 'priority', 'status', 'due_date', 'is_overdue']
    list_filter = ['priority', 'status', 'department', 'created_at']
    list_select_related = ['company', 'assigned_to', 'created_by']
    search_fields = ['title', 'description', 'company__name']
    readonly_fields = ['created_at', 'updated_at', 'completed_at']
    actions = ['bulk_mark_completed', 'bulk_mark_in_progress', 'bulk_assign_to_user']
    
    def is_overdue(self, obj):
        return obj.is_overdue
    is_overdue.boolean = True

    def bulk_mark_completed(self, request, queryset):
        """Mark selected tasks as completed"""
        updated_count = 0
        for task in queryset:
            if task.status != 'Completed':
                task.status = 'Completed'
                task.completed_at = timezone.now()
                task.save()
                updated_count += 1

        if updated_count > 0:
            self.message_user(request, f'Successfully marked {updated_count} tasks as completed')
        else:
            self.message_user(request, 'No tasks were updated (already completed)', level='WARNING')
    bulk_mark_completed.short_description = 'Mark selected tasks as completed'

    def bulk_mark_in_progress(self, request, queryset):
        """Mark selected tasks as in progress"""
        updated_count = queryset.exclude(status='In Progress').update(status='In Progress', completed_at=None)
        self.message_user(request, f'Successfully marked {updated_count} tasks as in progress')
    bulk_mark_in_progress.short_description = 'Mark selected tasks as in progress'

    def bulk_assign_to_user(self, request, queryset):
        """Bulk assign tasks to a user - redirect to form"""
        if queryset.count() > 50:
            self.message_user(request, 'Please select 50 tasks or fewer for bulk assignment', level='ERROR')
            return

        task_ids = ','.join(str(t.pk) for t in queryset)
        return HttpResponseRedirect(
            reverse('admin:crm_app_task_changelist') + f'?task_ids={task_ids}&action=assign'
        )
    bulk_assign_to_user.short_description = 'Bulk assign selected tasks to user'

    # list_select_related handles the optimization


class OpportunityActivityInline(admin.TabularInline):
    model = OpportunityActivity
    extra = 0
    fields = ['activity_type', 'subject', 'user', 'contact', 'created_at']
    readonly_fields = ['created_at']

    def get_queryset(self, request):
        """Optimize inline queries"""
        return super().get_queryset(request).select_related('opportunity', 'user', 'contact')


@admin.register(Opportunity)
class OpportunityAdmin(admin.ModelAdmin):
    list_display = ['name', 'company', 'stage', 'service_type', 'expected_value', 'probability', 'owner', 'expected_close_date']
    list_filter = ['stage', 'service_type', 'lead_source', 'is_active', 'created_at']
    list_select_related = ['company', 'owner']
    search_fields = ['name', 'company__name', 'notes']
    readonly_fields = ['created_at', 'updated_at', 'weighted_value', 'days_in_stage']
    inlines = [OpportunityActivityInline]
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('company', 'name', 'service_type', 'stage', 'owner', 'is_active')
        }),
        ('Financial', {
            'fields': ('expected_value', 'probability', 'weighted_value')
        }),
        ('Source & Method', {
            'fields': ('lead_source', 'contact_method')
        }),
        ('Dates', {
            'fields': ('last_contact_date', 'follow_up_date', 'expected_close_date', 'actual_close_date')
        }),
        ('Details', {
            'fields': ('competitors', 'pain_points', 'decision_criteria', 'notes')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at', 'days_in_stage'),
            'classes': ('collapse',)
        }),
    )


@admin.register(OpportunityActivity)
class OpportunityActivityAdmin(admin.ModelAdmin):
    list_display = ['subject', 'opportunity', 'activity_type', 'user', 'created_at']
    list_filter = ['activity_type', 'created_at']
    list_select_related = ['opportunity', 'opportunity__company', 'user', 'contact']
    search_fields = ['subject', 'description', 'opportunity__name']
    readonly_fields = ['created_at', 'updated_at']


class InvoiceInline(admin.TabularInline):
    model = Invoice
    extra = 0
    fields = ['invoice_number', 'status', 'issue_date', 'due_date', 'total_amount']

    def get_queryset(self, request):
        """Optimize inline queries"""
        return super().get_queryset(request).select_related('contract')


class ContractZoneInline(admin.TabularInline):
    """Inline admin for zones linked to a contract"""
    model = ContractZone
    extra = 0
    fields = ('zone', 'start_date', 'end_date', 'is_active', 'notes')
    readonly_fields = ('zone',)  # Can't change zone after creation

    def get_queryset(self, request):
        """Optimize inline queries"""
        return super().get_queryset(request).select_related('zone', 'zone__company')

    def has_delete_permission(self, request, obj=None):
        # Allow deletion only for inactive relationships
        return False


class ContractDocumentInline(admin.TabularInline):
    """Inline admin for contract documents"""
    model = ContractDocument
    extra = 0
    readonly_fields = ['uploaded_at', 'uploaded_by']
    fields = ['document_type', 'title', 'file', 'is_official', 'is_signed', 'signed_date', 'uploaded_at']

    def get_queryset(self, request):
        """Optimize inline queries"""
        return super().get_queryset(request).select_related('contract', 'uploaded_by')


class ContractAdminForm(forms.ModelForm):
    """Custom form for Contract admin with better number formatting"""
    class Meta:
        model = Contract
        fields = '__all__'
        localized_fields = ['value']  # This will apply number formatting


@admin.register(Contract)
class ContractAdmin(admin.ModelAdmin):
    form = ContractAdminForm
    list_display = ['contract_number', 'company', 'service_type', 'contract_type', 'contract_category_badge', 'status', 'start_date', 'end_date', 'value', 'is_expiring_soon', 'participation_count']
    list_filter = ['service_type', 'contract_type', 'contract_category', 'status', 'auto_renew', 'is_active']
    list_select_related = ['company', 'opportunity', 'master_contract', 'master_contract__company']
    search_fields = ['contract_number', 'company__name']
    readonly_fields = ['created_at', 'updated_at', 'days_until_expiry', 'formatted_monthly_value', 'participation_count']
    autocomplete_fields = ['master_contract']
    inlines = [InvoiceInline, ContractZoneInline, ContractDocumentInline]
    actions = ['bulk_update_status_active', 'bulk_update_status_inactive', 'export_contracts_csv', 'export_contracts_excel']
    
    def is_expiring_soon(self, obj):
        return obj.is_expiring_soon
    is_expiring_soon.boolean = True

    def contract_category_badge(self, obj):
        """Display colored badge for contract category"""
        color_map = {
            'standard': '#2196F3',  # Blue
            'corporate_master': '#FFA500',  # Orange
            'participation': '#4CAF50',  # Green
        }
        label_map = {
            'standard': 'STANDARD',
            'corporate_master': 'MASTER',
            'participation': 'PARTICIPATION',
        }
        color = color_map.get(obj.contract_category, '#999')
        label = label_map.get(obj.contract_category, obj.get_contract_category_display())

        # Add master contract reference for participation agreements
        extra_info = ''
        if obj.contract_category == 'participation' and obj.master_contract:
            extra_info = f'<br><small style="color: #666;">→ {obj.master_contract.contract_number}</small>'

        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; '
            'border-radius: 3px; font-size: 11px; font-weight: bold;">{}</span>{}',
            color, label, extra_info
        )
    contract_category_badge.short_description = "Category"

    def participation_count(self, obj):
        """Display count of participation agreements for master contracts"""
        if obj.contract_category == 'corporate_master':
            count = obj.participation_agreements.count()
            if count > 0:
                return format_html(
                    '<span style="background-color: #4CAF50; color: white; padding: 2px 6px; '
                    'border-radius: 3px; font-size: 11px; font-weight: bold;">{} agreements</span>',
                    count
                )
            return format_html('<span style="color: #999;">0 agreements</span>')
        return format_html('<span style="color: #999;">-</span>')
    participation_count.short_description = "Participation Agreements"

    def formatted_monthly_value(self, obj):
        """Display monthly value with proper currency formatting"""
        try:
            # Make sure we have the object
            if not obj or not obj.pk:
                return format_html('<span style="color: #999;">-</span>')
            
            # Check if we have required fields
            if not obj.start_date:
                return format_html('<span style="color: #999;">No start date</span>')
            if not obj.end_date:
                return format_html('<span style="color: #999;">No end date</span>')
            if not obj.value:
                return format_html('<span style="color: #999;">No value set</span>')
            
            # Calculate monthly value manually here to debug
            months = ((obj.end_date.year - obj.start_date.year) * 12 + 
                     (obj.end_date.month - obj.start_date.month))
            if obj.end_date.day >= obj.start_date.day:
                months += 1
            
            if months > 0:
                monthly = round(float(obj.value) / months, 2)
                # Format with thousands separator and 2 decimal places
                formatted_value = "{:,.2f}".format(monthly)
                return format_html(
                    '<span style="font-weight: normal;">{}</span>',
                    formatted_value
                )
            else:
                return format_html('<span style="color: #999;">Invalid date range</span>')
                
        except Exception as e:
            return format_html('<span style="color: red;">Error: {}</span>', str(e))
    formatted_monthly_value.short_description = 'Monthly value'
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('company', 'opportunity', 'contract_number', 'contract_type', 'service_type', 'status')
        }),
        ('Contract Category', {
            'fields': ('contract_category', 'master_contract', 'participation_count'),
            'description': 'Configure contract category and master agreement relationships for corporate contracts'
        }),
        ('Dates', {
            'fields': ('start_date', 'end_date', 'days_until_expiry')
        }),
        ('Financial', {
            'fields': ('value', 'formatted_monthly_value', 'currency', 'discount_percentage')
        }),
        ('Terms', {
            'fields': ('payment_terms', 'billing_frequency', 'auto_renew', 'renewal_period_months')
        }),
        ('Signatories', {
            'fields': ('customer_signatory_name', 'customer_signatory_title', 'bmasia_signatory_name', 'bmasia_signatory_title'),
            'classes': ('collapse',),
            'description': 'Authorized signatories for this contract'
        }),
        ('Custom Terms', {
            'fields': ('custom_terms',),
            'classes': ('collapse',),
            'description': 'Custom terms and conditions specific to this master agreement'
        }),
        ('Renewal Tracking', {
            'fields': ('renewal_notice_sent', 'renewal_notice_date'),
            'classes': ('collapse',)
        }),
        ('Other', {
            'fields': ('is_active', 'notes')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def get_form(self, request, obj=None, **kwargs):
        """Conditionally show master_contract field based on contract_category"""
        form = super().get_form(request, obj, **kwargs)

        # If this is a participation agreement, master_contract is required
        # For other categories, hide the field with help text
        if obj and obj.contract_category != 'participation':
            if 'master_contract' in form.base_fields:
                form.base_fields['master_contract'].widget = forms.HiddenInput()
                form.base_fields['master_contract'].required = False

        return form

    def bulk_update_status_active(self, request, queryset):
        """Mark selected contracts as active"""
        updated_count = queryset.update(status='Active', is_active=True)
        self.message_user(request, f'Successfully marked {updated_count} contracts as active')
    bulk_update_status_active.short_description = 'Mark selected contracts as active'

    def bulk_update_status_inactive(self, request, queryset):
        """Mark selected contracts as inactive"""
        updated_count = queryset.update(status='Terminated', is_active=False)
        self.message_user(request, f'Successfully marked {updated_count} contracts as inactive')
    bulk_update_status_inactive.short_description = 'Mark selected contracts as inactive'

    def export_contracts_csv(self, request, queryset):
        """Export selected contracts to CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="contracts_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'

        writer = csv.writer(response)
        # Write headers
        writer.writerow([
            'Contract Number', 'Company', 'Service Type', 'Contract Type', 'Status',
            'Start Date', 'End Date', 'Value', 'Currency', 'Monthly Value',
            'Payment Terms', 'Billing Frequency', 'Auto Renew', 'Is Active',
            'Days Until Expiry', 'Created Date', 'Last Updated'
        ])

        # Write data with optimized queries
        contracts = queryset.select_related('company', 'opportunity')

        for contract in contracts:
            writer.writerow([
                contract.contract_number,
                contract.company.name,
                contract.get_service_type_display() if contract.service_type else '',
                contract.get_contract_type_display(),
                contract.get_status_display(),
                contract.start_date.strftime('%Y-%m-%d'),
                contract.end_date.strftime('%Y-%m-%d'),
                f"{contract.value:.2f}",
                contract.currency,
                f"{contract.monthly_value:.2f}",
                contract.payment_terms,
                contract.billing_frequency,
                'Yes' if contract.auto_renew else 'No',
                'Yes' if contract.is_active else 'No',
                contract.days_until_expiry,
                contract.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                contract.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])

        self.message_user(request, f'Successfully exported {queryset.count()} contracts to CSV')
        return response
    export_contracts_csv.short_description = 'Export selected contracts to CSV'

    def export_contracts_excel(self, request, queryset):
        """Export selected contracts to Excel with formatting"""
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contracts Export"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        currency_format = '#,##0.00'

        # Write headers
        headers = [
            'Contract Number', 'Company', 'Service Type', 'Contract Type', 'Status',
            'Start Date', 'End Date', 'Value', 'Currency', 'Monthly Value',
            'Payment Terms', 'Billing Frequency', 'Auto Renew', 'Is Active',
            'Days Until Expiry', 'Created Date', 'Last Updated'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data with optimized queries
        contracts = queryset.select_related('company', 'opportunity')

        for row, contract in enumerate(contracts, 2):
            ws.cell(row=row, column=1, value=contract.contract_number)
            ws.cell(row=row, column=2, value=contract.company.name)
            ws.cell(row=row, column=3, value=contract.get_service_type_display() if contract.service_type else '')
            ws.cell(row=row, column=4, value=contract.get_contract_type_display())
            ws.cell(row=row, column=5, value=contract.get_status_display())
            ws.cell(row=row, column=6, value=contract.start_date)
            ws.cell(row=row, column=7, value=contract.end_date)

            # Format currency values
            value_cell = ws.cell(row=row, column=8, value=float(contract.value))
            value_cell.number_format = currency_format

            ws.cell(row=row, column=9, value=contract.currency)

            monthly_cell = ws.cell(row=row, column=10, value=float(contract.monthly_value))
            monthly_cell.number_format = currency_format

            ws.cell(row=row, column=11, value=contract.payment_terms)
            ws.cell(row=row, column=12, value=contract.billing_frequency)
            ws.cell(row=row, column=13, value='Yes' if contract.auto_renew else 'No')
            ws.cell(row=row, column=14, value='Yes' if contract.is_active else 'No')
            ws.cell(row=row, column=15, value=contract.days_until_expiry)
            ws.cell(row=row, column=16, value=contract.created_at)
            ws.cell(row=row, column=17, value=contract.updated_at)

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = max(len(str(ws.cell(row=1, column=col).value)), 15)
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="contracts_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        wb.save(response)
        self.message_user(request, f'Successfully exported {queryset.count()} contracts to Excel')
        return response
    export_contracts_excel.short_description = 'Export selected contracts to Excel'

    # list_select_related handles the optimization


@admin.register(ContractZone)
class ContractZoneAdmin(admin.ModelAdmin):
    """Admin interface for ContractZone - historical tracking of zone-contract relationships"""
    list_display = ('contract', 'zone', 'start_date', 'end_date', 'is_active', 'created_at')
    list_filter = ('is_active', 'start_date', 'end_date')
    list_select_related = ('contract', 'contract__company', 'zone', 'zone__company')
    search_fields = ('contract__contract_number', 'zone__name', 'zone__company__name')
    date_hierarchy = 'start_date'
    readonly_fields = ('created_at', 'updated_at')

    fieldsets = (
        ('Relationship', {
            'fields': ('contract', 'zone')
        }),
        ('Dates', {
            'fields': ('start_date', 'end_date', 'is_active')
        }),
        ('Additional Information', {
            'fields': ('notes', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def get_queryset(self, request):
        """Optimize queries with select_related"""
        return super().get_queryset(request).select_related(
            'contract',
            'contract__company',
            'zone',
            'zone__company'
        )


@admin.register(Invoice)
class InvoiceAdmin(admin.ModelAdmin):
    list_display = ['invoice_number', 'contract', 'status', 'issue_date', 'due_date', 'total_amount', 'days_overdue']
    list_filter = ['status', 'issue_date', 'due_date']
    list_select_related = ['contract', 'contract__company']
    search_fields = ['invoice_number', 'contract__company__name', 'contract__contract_number']
    readonly_fields = ['created_at', 'updated_at', 'total_amount', 'days_overdue']
    actions = ['bulk_mark_paid', 'bulk_mark_unpaid', 'export_invoices_csv', 'export_invoices_excel']
    
    def days_overdue(self, obj):
        return obj.days_overdue if obj.is_overdue else 0
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('contract', 'invoice_number', 'status')
        }),
        ('Dates', {
            'fields': ('issue_date', 'due_date', 'paid_date', 'days_overdue')
        }),
        ('Financial', {
            'fields': ('amount', 'tax_amount', 'discount_amount', 'total_amount', 'currency')
        }),
        ('Payment', {
            'fields': ('payment_method', 'transaction_id')
        }),
        ('Reminders', {
            'fields': ('first_reminder_sent', 'second_reminder_sent', 'final_notice_sent'),
            'classes': ('collapse',)
        }),
        ('Other', {
            'fields': ('notes',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def bulk_mark_paid(self, request, queryset):
        """Mark selected invoices as paid"""
        updated_count = 0
        for invoice in queryset:
            if invoice.status != 'Paid':
                invoice.status = 'Paid'
                invoice.paid_date = timezone.now().date()
                invoice.save()
                updated_count += 1

        if updated_count > 0:
            self.message_user(request, f'Successfully marked {updated_count} invoices as paid')
        else:
            self.message_user(request, 'No invoices were updated (already paid)', level='WARNING')
    bulk_mark_paid.short_description = 'Mark selected invoices as paid'

    def bulk_mark_unpaid(self, request, queryset):
        """Mark selected invoices as unpaid"""
        updated_count = queryset.exclude(status='Sent').update(status='Sent', paid_date=None)
        self.message_user(request, f'Successfully marked {updated_count} invoices as unpaid')
    bulk_mark_unpaid.short_description = 'Mark selected invoices as unpaid'

    def export_invoices_csv(self, request, queryset):
        """Export selected invoices to CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="invoices_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'

        writer = csv.writer(response)
        # Write headers
        writer.writerow([
            'Invoice Number', 'Contract Number', 'Company', 'Status', 'Issue Date',
            'Due Date', 'Paid Date', 'Amount', 'Tax Amount', 'Discount Amount',
            'Total Amount', 'Currency', 'Payment Method', 'Days Overdue',
            'Transaction ID', 'Notes', 'Created Date', 'Last Updated'
        ])

        # Write data with optimized queries
        invoices = queryset.select_related('contract', 'contract__company')

        for invoice in invoices:
            writer.writerow([
                invoice.invoice_number,
                invoice.contract.contract_number,
                invoice.contract.company.name,
                invoice.get_status_display(),
                invoice.issue_date.strftime('%Y-%m-%d'),
                invoice.due_date.strftime('%Y-%m-%d'),
                invoice.paid_date.strftime('%Y-%m-%d') if invoice.paid_date else '',
                f"{invoice.amount:.2f}",
                f"{invoice.tax_amount:.2f}",
                f"{invoice.discount_amount:.2f}",
                f"{invoice.total_amount:.2f}",
                invoice.currency,
                invoice.payment_method,
                invoice.days_overdue,
                invoice.transaction_id,
                invoice.notes,
                invoice.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                invoice.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])

        self.message_user(request, f'Successfully exported {queryset.count()} invoices to CSV')
        return response
    export_invoices_csv.short_description = 'Export selected invoices to CSV'

    def export_invoices_excel(self, request, queryset):
        """Export selected invoices to Excel with formatting"""
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoices Export"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        currency_format = '#,##0.00'

        # Write headers
        headers = [
            'Invoice Number', 'Contract Number', 'Company', 'Status', 'Issue Date',
            'Due Date', 'Paid Date', 'Amount', 'Tax Amount', 'Discount Amount',
            'Total Amount', 'Currency', 'Payment Method', 'Days Overdue',
            'Transaction ID', 'Notes', 'Created Date', 'Last Updated'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data with optimized queries
        invoices = queryset.select_related('contract', 'contract__company')

        for row, invoice in enumerate(invoices, 2):
            ws.cell(row=row, column=1, value=invoice.invoice_number)
            ws.cell(row=row, column=2, value=invoice.contract.contract_number)
            ws.cell(row=row, column=3, value=invoice.contract.company.name)
            ws.cell(row=row, column=4, value=invoice.get_status_display())
            ws.cell(row=row, column=5, value=invoice.issue_date)
            ws.cell(row=row, column=6, value=invoice.due_date)
            ws.cell(row=row, column=7, value=invoice.paid_date)

            # Format currency values
            amount_cell = ws.cell(row=row, column=8, value=float(invoice.amount))
            amount_cell.number_format = currency_format

            tax_cell = ws.cell(row=row, column=9, value=float(invoice.tax_amount))
            tax_cell.number_format = currency_format

            discount_cell = ws.cell(row=row, column=10, value=float(invoice.discount_amount))
            discount_cell.number_format = currency_format

            total_cell = ws.cell(row=row, column=11, value=float(invoice.total_amount))
            total_cell.number_format = currency_format

            ws.cell(row=row, column=12, value=invoice.currency)
            ws.cell(row=row, column=13, value=invoice.payment_method)
            ws.cell(row=row, column=14, value=invoice.days_overdue)
            ws.cell(row=row, column=15, value=invoice.transaction_id)
            ws.cell(row=row, column=16, value=invoice.notes)
            ws.cell(row=row, column=17, value=invoice.created_at)
            ws.cell(row=row, column=18, value=invoice.updated_at)

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = max(len(str(ws.cell(row=1, column=col).value)), 15)
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="invoices_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        wb.save(response)
        self.message_user(request, f'Successfully exported {queryset.count()} invoices to Excel')
        return response
    export_invoices_excel.short_description = 'Export selected invoices to Excel'

    # list_select_related handles the optimization


@admin.register(AuditLog)
class AuditLogAdmin(admin.ModelAdmin):
    list_display = ['user', 'action', 'model_name', 'record_id', 'timestamp']
    list_filter = ['action', 'model_name', 'timestamp']
    search_fields = ['user__username', 'model_name', 'record_id']
    readonly_fields = ['user', 'action', 'model_name', 'record_id', 'timestamp', 'ip_address', 'user_agent', 'changes', 'additional_data']
    
    def has_add_permission(self, request):
        return False
    
    def has_change_permission(self, request, obj=None):
        return False


# DEPRECATED: SubscriptionPlan removed - use Contract model with service_type instead
# See admin_backup_subscription.py for the old code


@admin.register(Device)
class DeviceAdmin(admin.ModelAdmin):
    list_display = ['name', 'device_type', 'company', 'zone_count', 'created_at']
    list_filter = ['device_type', 'company']
    search_fields = ['name', 'model_info', 'notes', 'company__name']
    readonly_fields = ['id', 'created_at', 'updated_at', 'zone_count']
    autocomplete_fields = ['company']

    fieldsets = (
        (None, {
            'fields': ('company', 'name', 'device_type')
        }),
        ('Details', {
            'fields': ('model_info', 'notes'),
            'classes': ('collapse',)
        }),
        ('Metadata', {
            'fields': ('id', 'zone_count', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def zone_count(self, obj):
        return obj.zones.count()
    zone_count.short_description = 'Zones'


@admin.register(Zone)
class ZoneAdmin(admin.ModelAdmin):
    list_display = ["zone_name_with_company", "platform", "currently_playing_admin", "status_badge", "last_seen_online", "last_api_sync"]
    list_filter = ["platform", "status", "company"]
    list_select_related = ["company"]
    search_fields = ["name", "company__name", "company__soundtrack_account_id"]
    readonly_fields = ["created_at", "updated_at", "last_api_sync", "api_raw_data", "status_badge", "soundtrack_account_id", "currently_playing_admin"]
    
    def zone_name_with_company(self, obj):
        return f"{obj.company.name} - {obj.name}"
    zone_name_with_company.short_description = "Zone"
    zone_name_with_company.admin_order_field = "company__name"
    
    def currently_playing_admin(self, obj):
        """Display what's currently playing"""
        if obj.api_raw_data and obj.api_raw_data.get('currently_playing'):
            return obj.api_raw_data.get('currently_playing')
        return "No active playlist/schedule"
    currently_playing_admin.short_description = "Currently Playing"
    
    def status_badge(self, obj):
        colors = {
            "online": "green",
            "offline": "orange",
            "no_device": "red",
            "expired": "gray",
            "pending": "blue"
        }
        color = colors.get(obj.status, "gray")
        return format_html(
            "<span style=\"padding: 3px 10px; border-radius: 3px; color: white; background-color: {};\">{}</span>",
            color, obj.get_status_display()
        )
    status_badge.short_description = "Status"
    
    fieldsets = (
        ("Zone Information", {
            "fields": ("company", "name", "platform")
        }),
        ("Status", {
            "fields": ("status", "status_badge", "currently_playing_admin", "last_seen_online", "device_name")
        }),
        ("Soundtrack Integration", {
            "fields": ("soundtrack_account_id", "soundtrack_zone_id", "soundtrack_admin_email"),
            "description": "Account ID is inherited from the company. Zone-specific data from API is shown here."
        }),
        ("API Sync", {
            "fields": ("last_api_sync", "api_raw_data"),
            "classes": ("collapse",)
        }),
        ("Additional Info", {
            "fields": ("notes",)
        }),
        ("Timestamps", {
            "fields": ("created_at", "updated_at"),
            "classes": ("collapse",)
        }),
    )
    
    actions = ["sync_with_soundtrack_api"]
    
    def sync_with_soundtrack_api(self, request, queryset):
        """Action to sync selected zones with Soundtrack API"""
        from django.utils import timezone
        from .services.soundtrack_api import soundtrack_api
        
        # Group zones by company for efficient syncing
        companies_to_sync = set()
        for zone in queryset.filter(platform="soundtrack"):
            if zone.company.soundtrack_account_id:
                companies_to_sync.add(zone.company)
        
        total_synced = 0
        for company in companies_to_sync:
            try:
                synced, errors = soundtrack_api.sync_company_zones(company)
                total_synced += synced
                if synced > 0:
                    self.message_user(request, f"✓ Synced {synced} zones for {company.name}")
            except Exception as e:
                self.message_user(request, f"Error syncing {company.name}: {str(e)}", level="ERROR")
        
        if total_synced > 0:
            self.message_user(request, f"Successfully synced {total_synced} zones total")
    sync_with_soundtrack_api.short_description = "Sync with Soundtrack API"


@admin.register(ZoneOfflineAlert)
class ZoneOfflineAlertAdmin(admin.ModelAdmin):
    list_display = ['zone_name', 'company_name', 'status_badge', 'hours_offline_display', 'notification_count', 'detected_at']
    list_filter = ['is_resolved', 'zone__company']
    list_select_related = ['zone', 'zone__company']
    search_fields = ['zone__name', 'zone__company__name']
    readonly_fields = ['detected_at', 'resolved_at', 'first_notification_at', 'last_notification_at', 'hours_offline_display']
    ordering = ['-detected_at']

    def zone_name(self, obj):
        return obj.zone.name
    zone_name.short_description = "Zone"
    zone_name.admin_order_field = "zone__name"

    def company_name(self, obj):
        return obj.zone.company.name
    company_name.short_description = "Company"
    company_name.admin_order_field = "zone__company__name"

    def status_badge(self, obj):
        if obj.is_resolved:
            return format_html('<span style="padding: 3px 10px; border-radius: 3px; color: white; background-color: green;">Resolved</span>')
        return format_html('<span style="padding: 3px 10px; border-radius: 3px; color: white; background-color: red;">Active</span>')
    status_badge.short_description = "Status"

    def hours_offline_display(self, obj):
        hours = obj.hours_offline
        if hours < 1:
            return f"{int(hours * 60)} minutes"
        elif hours < 24:
            return f"{hours:.1f} hours"
        else:
            days = hours / 24
            return f"{days:.1f} days"
    hours_offline_display.short_description = "Duration"


# Email System Admin
@admin.register(EmailTemplate)
class EmailTemplateAdmin(admin.ModelAdmin):
    list_display = ['name', 'template_type', 'language', 'department', 'is_active']
    list_filter = ['template_type', 'language', 'department', 'is_active']
    list_select_related = []  # EmailTemplate has no ForeignKey fields to select_related on
    search_fields = ['name', 'subject', 'body_text']
    readonly_fields = ['created_at', 'updated_at']
    
    fieldsets = (
        ('Template Info', {
            'fields': ('name', 'template_type', 'language', 'department', 'is_active')
        }),
        ('Email Content', {
            'fields': ('subject', 'body_text'),
            'description': 'Write your email in plain text. HTML will be generated automatically. Use variables like {{company_name}}, {{contact_name}}, {{days_until_expiry}}'
        }),
        ('Advanced', {
            'fields': ('body_html',),
            'classes': ('collapse',),
            'description': 'Optional: Custom HTML template (leave empty to auto-generate from text)'
        }),
        ('Notes', {
            'fields': ('notes',),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    actions = ['preview_template', 'send_email', 'duplicate_template']

    def preview_template(self, request, queryset):
        """Preview selected template"""
        if queryset.count() != 1:
            self.message_user(request, 'Please select exactly one template', level='ERROR')
            return

        template = queryset.first()
        # Redirect to preview template
        from django.urls import reverse
        from django.http import HttpResponseRedirect
        return HttpResponseRedirect(
            reverse('admin_preview_template', args=[template.pk])
        )
    preview_template.short_description = 'Preview template with sample data'

    def send_email(self, request, queryset):
        """Send email using selected template"""
        if queryset.count() != 1:
            self.message_user(request, 'Please select exactly one template', level='ERROR')
            return

        template = queryset.first()
        # Redirect to send email form
        from django.urls import reverse
        from django.http import HttpResponseRedirect
        return HttpResponseRedirect(
            reverse('admin_send_email_template', args=[template.pk])
        )
    send_email.short_description = 'Send email using this template'
    
    def duplicate_template(self, request, queryset):
        """Duplicate selected templates"""
        for template in queryset:
            template.pk = None
            template.name = f"{template.name} (Copy)"
            template.save()
        self.message_user(request, f'Duplicated {queryset.count()} template(s)')
    duplicate_template.short_description = 'Duplicate selected templates'


@admin.register(EmailLog)
class EmailLogAdmin(admin.ModelAdmin):
    list_display = ['subject', 'to_email', 'email_type', 'status', 'sent_at', 'company']
    list_filter = ['status', 'email_type', 'created_at', 'sent_at']
    list_select_related = ['company', 'contact', 'template_used', 'contract', 'invoice']
    search_fields = ['subject', 'to_email', 'from_email', 'company__name', 'contact__name']
    readonly_fields = [
        'created_at', 'updated_at', 'sent_at', 'opened_at', 'clicked_at',
        'message_id', 'status_badge', 'attachments'
    ]
    date_hierarchy = 'created_at'
    
    def status_badge(self, obj):
        colors = {
            'pending': '#ffc107',
            'sent': '#28a745',
            'failed': '#dc3545',
            'bounced': '#dc3545',
            'opened': '#17a2b8',
            'clicked': '#007bff',
            'unsubscribed': '#6c757d',
        }
        color = colors.get(obj.status, '#6c757d')
        return format_html(
            '<span style="padding: 3px 10px; border-radius: 3px; color: white; background-color: {};">{}</span>',
            color, obj.get_status_display()
        )
    status_badge.short_description = 'Status'
    
    fieldsets = (
        ('Email Details', {
            'fields': ('email_type', 'template_used', 'from_email', 'to_email', 'cc_emails', 'subject')
        }),
        ('Recipients', {
            'fields': ('company', 'contact', 'contract', 'invoice')
        }),
        ('Status', {
            'fields': ('status', 'status_badge', 'sent_at', 'opened_at', 'clicked_at', 'error_message')
        }),
        ('Attachments', {
            'fields': ('attachments',),
            'description': 'Documents attached to this email'
        }),
        ('Content', {
            'fields': ('body_html', 'body_text'),
            'classes': ('collapse',)
        }),
        ('Tracking', {
            'fields': ('message_id', 'in_reply_to'),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def has_add_permission(self, request):
        return False  # Emails should only be created by the system
    
    def has_change_permission(self, request, obj=None):
        return False  # Email logs should not be edited


class CampaignRecipientInline(admin.TabularInline):
    """Inline admin for campaign recipients - show first 10"""
    model = CampaignRecipient
    extra = 0
    max_num = 10
    can_delete = False
    fields = ['contact', 'contact_email', 'status', 'sent_at', 'opened_at', 'clicked_at']
    readonly_fields = ['contact', 'contact_email', 'status', 'sent_at', 'opened_at', 'clicked_at']

    def contact_email(self, obj):
        return obj.contact.email if obj.contact else '-'
    contact_email.short_description = 'Email'

    def has_add_permission(self, request, obj):
        return False


@admin.register(EmailCampaign)
class EmailCampaignAdmin(admin.ModelAdmin):
    list_display = [
        'name', 'campaign_type', 'status', 'audience_count',
        'scheduled_send_date', 'total_sent', 'open_rate_display',
        'click_rate_display', 'bounce_rate_display'
    ]
    list_filter = ['status', 'campaign_type', 'created_at', 'scheduled_send_date']
    list_select_related = ['company', 'contract', 'template']
    search_fields = ['name', 'subject', 'company__name']
    readonly_fields = [
        'created_at', 'updated_at', 'actual_send_date', 'audience_count',
        'total_sent', 'total_delivered', 'total_bounced', 'total_opened',
        'total_clicked', 'total_unsubscribed', 'total_complained',
        'open_rate_display', 'click_rate_display', 'bounce_rate_display',
        'emails_sent', 'last_email_sent'  # Legacy fields
    ]
    inlines = [CampaignRecipientInline]
    date_hierarchy = 'created_at'
    actions = [
        'mark_as_sent', 'pause_campaigns', 'resume_campaigns',
        'cancel_campaigns', 'export_campaign_analytics'
    ]

    fieldsets = (
        ('Campaign Information', {
            'fields': ('name', 'campaign_type', 'subject', 'template', 'status')
        }),
        ('Targeting', {
            'fields': ('target_audience', 'audience_count'),
            'description': 'Define audience segmentation criteria (JSON format). Leave blank to manually add recipients.'
        }),
        ('Scheduling', {
            'fields': ('scheduled_send_date', 'actual_send_date', 'send_immediately')
        }),
        ('Email Settings', {
            'fields': ('sender_email', 'reply_to_email'),
            'description': 'Leave sender_email blank to use default system email'
        }),
        ('Analytics', {
            'fields': (
                ('total_sent', 'total_delivered', 'total_bounced'),
                ('total_opened', 'total_clicked', 'total_unsubscribed'),
                ('open_rate_display', 'click_rate_display', 'bounce_rate_display'),
            ),
            'description': 'Campaign performance metrics (auto-updated)'
        }),
        ('Legacy Fields (Backward Compatibility)', {
            'fields': (
                'company', 'contract', 'is_active', 'start_date', 'end_date',
                'emails_sent', 'last_email_sent', 'stop_on_reply', 'replied'
            ),
            'classes': ('collapse',),
            'description': 'These fields are kept for backward compatibility with older campaigns'
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def open_rate_display(self, obj):
        """Display open rate with color coding"""
        rate = obj.open_rate
        if rate >= 20:
            color = 'green'
        elif rate >= 10:
            color = 'orange'
        else:
            color = 'red'

        return format_html(
            '<span style="color: {}; font-weight: bold;">{:.1f}%</span>',
            color, rate
        )
    open_rate_display.short_description = 'Open Rate'

    def click_rate_display(self, obj):
        """Display click rate with color coding"""
        rate = obj.click_rate
        if rate >= 3:
            color = 'green'
        elif rate >= 1:
            color = 'orange'
        else:
            color = 'red'

        return format_html(
            '<span style="color: {}; font-weight: bold;">{:.1f}%</span>',
            color, rate
        )
    click_rate_display.short_description = 'Click Rate'

    def bounce_rate_display(self, obj):
        """Display bounce rate with color coding"""
        rate = obj.bounce_rate
        if rate <= 2:
            color = 'green'
        elif rate <= 5:
            color = 'orange'
        else:
            color = 'red'

        return format_html(
            '<span style="color: {}; font-weight: bold;">{:.1f}%</span>',
            color, rate
        )
    bounce_rate_display.short_description = 'Bounce Rate'

    def mark_as_sent(self, request, queryset):
        """Mark selected campaigns as sent"""
        count = 0
        for campaign in queryset:
            if campaign.status in ['draft', 'scheduled']:
                campaign.status = 'sent'
                campaign.actual_send_date = timezone.now()
                campaign.save()
                count += 1

        if count > 0:
            self.message_user(request, f'Marked {count} campaign(s) as sent')
        else:
            self.message_user(request, 'No campaigns were updated (already sent or in progress)', level='WARNING')
    mark_as_sent.short_description = 'Mark selected campaigns as sent'

    def pause_campaigns(self, request, queryset):
        """Pause selected campaigns"""
        count = queryset.filter(status__in=['draft', 'scheduled', 'sending']).update(
            status='paused',
            is_active=False
        )
        self.message_user(request, f'Paused {count} campaign(s)')
    pause_campaigns.short_description = 'Pause selected campaigns'

    def resume_campaigns(self, request, queryset):
        """Resume paused campaigns"""
        count = queryset.filter(status='paused').update(
            status='draft',
            is_active=True
        )
        self.message_user(request, f'Resumed {count} campaign(s)')
    resume_campaigns.short_description = 'Resume selected campaigns'

    def cancel_campaigns(self, request, queryset):
        """Cancel selected campaigns"""
        count = queryset.exclude(status__in=['sent', 'cancelled']).update(
            status='cancelled',
            is_active=False
        )
        self.message_user(request, f'Cancelled {count} campaign(s)')
    cancel_campaigns.short_description = 'Cancel selected campaigns'

    def export_campaign_analytics(self, request, queryset):
        """Export campaign analytics to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Campaign Analytics"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        headers = [
            'Campaign Name', 'Type', 'Status', 'Subject', 'Scheduled Date',
            'Actual Send Date', 'Total Sent', 'Delivered', 'Bounced', 'Opened',
            'Clicked', 'Unsubscribed', 'Open Rate %', 'Click Rate %', 'Bounce Rate %',
            'Created Date'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data
        campaigns = queryset.select_related('company', 'template')

        for row, campaign in enumerate(campaigns, 2):
            ws.cell(row=row, column=1, value=campaign.name)
            ws.cell(row=row, column=2, value=campaign.get_campaign_type_display())
            ws.cell(row=row, column=3, value=campaign.get_status_display())
            ws.cell(row=row, column=4, value=campaign.subject)
            ws.cell(row=row, column=5, value=campaign.scheduled_send_date)
            ws.cell(row=row, column=6, value=campaign.actual_send_date)
            ws.cell(row=row, column=7, value=campaign.total_sent)
            ws.cell(row=row, column=8, value=campaign.total_delivered)
            ws.cell(row=row, column=9, value=campaign.total_bounced)
            ws.cell(row=row, column=10, value=campaign.total_opened)
            ws.cell(row=row, column=11, value=campaign.total_clicked)
            ws.cell(row=row, column=12, value=campaign.total_unsubscribed)
            ws.cell(row=row, column=13, value=campaign.open_rate)
            ws.cell(row=row, column=14, value=campaign.click_rate)
            ws.cell(row=row, column=15, value=campaign.bounce_rate)
            ws.cell(row=row, column=16, value=campaign.created_at)

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = max(len(str(ws.cell(row=1, column=col).value)), 15)
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="campaign_analytics_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        wb.save(response)
        self.message_user(request, f'Successfully exported {queryset.count()} campaigns to Excel')
        return response
    export_campaign_analytics.short_description = 'Export campaign analytics to Excel'

    def get_queryset(self, request):
        """Optimize queries"""
        return super().get_queryset(request).select_related('company', 'contract', 'template')


@admin.register(CampaignRecipient)
class CampaignRecipientAdmin(admin.ModelAdmin):
    list_display = [
        'campaign', 'contact', 'contact_email', 'status',
        'sent_at', 'opened_at', 'clicked_at', 'error_message_short'
    ]
    list_filter = ['status', 'sent_at', 'opened_at', 'clicked_at', 'bounced_at']
    list_select_related = ['campaign', 'contact', 'contact__company', 'email_log']
    search_fields = [
        'contact__email', 'contact__name',
        'contact__company__name', 'campaign__name', 'campaign__subject'
    ]
    readonly_fields = [
        'campaign', 'contact', 'email_log', 'status',
        'sent_at', 'delivered_at', 'opened_at', 'clicked_at',
        'bounced_at', 'failed_at', 'error_message',
        'created_at', 'updated_at'
    ]
    date_hierarchy = 'sent_at'
    actions = ['export_recipients_csv', 'export_recipients_excel']

    fieldsets = (
        ('Recipient Information', {
            'fields': ('campaign', 'contact', 'email_log')
        }),
        ('Status', {
            'fields': ('status', 'error_message')
        }),
        ('Timeline', {
            'fields': (
                ('sent_at', 'delivered_at'),
                ('opened_at', 'clicked_at'),
                ('bounced_at', 'failed_at')
            ),
            'description': 'Detailed timeline of email journey'
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def contact_email(self, obj):
        """Display contact email"""
        return obj.contact.email if obj.contact else '-'
    contact_email.short_description = 'Contact Email'
    contact_email.admin_order_field = 'contact__email'

    def error_message_short(self, obj):
        """Display truncated error message"""
        if obj.error_message:
            msg = obj.error_message[:50]
            if len(obj.error_message) > 50:
                msg += '...'
            return format_html('<span style="color: red;">{}</span>', msg)
        return '-'
    error_message_short.short_description = 'Error'

    def has_add_permission(self, request):
        """Prevent manual creation - recipients should be added programmatically"""
        return False

    def has_delete_permission(self, request, obj=None):
        """Allow deletion for cleanup"""
        return True

    def export_recipients_csv(self, request, queryset):
        """Export selected recipients to CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="campaign_recipients_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'

        writer = csv.writer(response)
        # Write headers
        writer.writerow([
            'Campaign Name', 'Campaign Subject', 'Contact Name', 'Contact Email',
            'Company', 'Status', 'Sent At', 'Delivered At', 'Opened At',
            'Clicked At', 'Bounced At', 'Failed At', 'Error Message'
        ])

        # Write data with optimized queries
        recipients = queryset.select_related('campaign', 'contact', 'contact__company')

        for recipient in recipients:
            writer.writerow([
                recipient.campaign.name,
                recipient.campaign.subject,
                recipient.contact.name,
                recipient.contact.email,
                recipient.contact.company.name,
                recipient.get_status_display(),
                recipient.sent_at.strftime('%Y-%m-%d %H:%M:%S') if recipient.sent_at else '',
                recipient.delivered_at.strftime('%Y-%m-%d %H:%M:%S') if recipient.delivered_at else '',
                recipient.opened_at.strftime('%Y-%m-%d %H:%M:%S') if recipient.opened_at else '',
                recipient.clicked_at.strftime('%Y-%m-%d %H:%M:%S') if recipient.clicked_at else '',
                recipient.bounced_at.strftime('%Y-%m-%d %H:%M:%S') if recipient.bounced_at else '',
                recipient.failed_at.strftime('%Y-%m-%d %H:%M:%S') if recipient.failed_at else '',
                recipient.error_message
            ])

        self.message_user(request, f'Successfully exported {queryset.count()} recipients to CSV')
        return response
    export_recipients_csv.short_description = 'Export selected recipients to CSV'

    def export_recipients_excel(self, request, queryset):
        """Export selected recipients to Excel with formatting"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Campaign Recipients"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        headers = [
            'Campaign Name', 'Campaign Subject', 'Contact Name', 'Contact Email',
            'Company', 'Status', 'Sent At', 'Delivered At', 'Opened At',
            'Clicked At', 'Bounced At', 'Failed At', 'Error Message'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data with optimized queries
        recipients = queryset.select_related('campaign', 'contact', 'contact__company')

        for row, recipient in enumerate(recipients, 2):
            ws.cell(row=row, column=1, value=recipient.campaign.name)
            ws.cell(row=row, column=2, value=recipient.campaign.subject)
            ws.cell(row=row, column=3, value=recipient.contact.name)
            ws.cell(row=row, column=4, value=recipient.contact.email)
            ws.cell(row=row, column=5, value=recipient.contact.company.name)
            ws.cell(row=row, column=6, value=recipient.get_status_display())
            ws.cell(row=row, column=7, value=recipient.sent_at)
            ws.cell(row=row, column=8, value=recipient.delivered_at)
            ws.cell(row=row, column=9, value=recipient.opened_at)
            ws.cell(row=row, column=10, value=recipient.clicked_at)
            ws.cell(row=row, column=11, value=recipient.bounced_at)
            ws.cell(row=row, column=12, value=recipient.failed_at)
            ws.cell(row=row, column=13, value=recipient.error_message)

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = max(len(str(ws.cell(row=1, column=col).value)), 15)
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="campaign_recipients_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        wb.save(response)
        self.message_user(request, f'Successfully exported {queryset.count()} recipients to Excel')
        return response
    export_recipients_excel.short_description = 'Export selected recipients to Excel'

    def get_queryset(self, request):
        """Optimize queries"""
        return super().get_queryset(request).select_related(
            'campaign', 'contact', 'contact__company', 'email_log'
        )


@admin.register(SeasonalTriggerDate)
class SeasonalTriggerDateAdmin(admin.ModelAdmin):
    list_display = ['holiday_type', 'year', 'trigger_date', 'holiday_date', 'updated_at', 'updated_by']
    list_filter = ['holiday_type', 'year']
    search_fields = ['notes']
    ordering = ['-year', 'trigger_date']
    readonly_fields = ['id', 'created_at', 'updated_at', 'updated_by']

    fieldsets = (
        ('Holiday Information', {
            'fields': ('holiday_type', 'year', 'holiday_date')
        }),
        ('Trigger Configuration', {
            'fields': ('trigger_date', 'notes'),
            'description': 'Set the date when campaign emails should be sent (typically 2 weeks before the holiday)'
        }),
        ('Metadata', {
            'fields': ('id', 'created_at', 'updated_at', 'updated_by'),
            'classes': ('collapse',)
        }),
    )

    def save_model(self, request, obj, form, change):
        obj.updated_by = request.user
        super().save_model(request, obj, form, change)

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('updated_by')


@admin.register(DocumentAttachment)
class DocumentAttachmentAdmin(admin.ModelAdmin):
    list_display = ['name', 'company', 'document_type', 'file', 'is_active', 'created_at']
    list_filter = ['document_type', 'is_active', 'created_at']
    list_select_related = ['company', 'contract', 'invoice']
    search_fields = ['name', 'description', 'company__name']
    readonly_fields = ['created_at', 'updated_at']

    fieldsets = (
        ('Document Info', {
            'fields': ('company', 'name', 'document_type', 'file', 'is_active')
        }),
        ('Related Objects', {
            'fields': ('contract', 'invoice'),
            'classes': ('collapse',)
        }),
        ('Details', {
            'fields': ('description',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )


# Quote Management Admin
class QuoteLineItemInline(admin.TabularInline):
    model = QuoteLineItem
    extra = 1
    fields = ['product_service', 'description', 'quantity', 'unit_price', 'discount_percentage', 'tax_rate', 'line_total']
    readonly_fields = ['line_total']


class QuoteActivityInline(admin.TabularInline):
    model = QuoteActivity
    extra = 0
    fields = ['activity_type', 'user', 'description', 'created_at']
    readonly_fields = ['created_at']


class QuoteAttachmentInline(admin.TabularInline):
    model = QuoteAttachment
    extra = 0
    fields = ['name', 'file', 'size', 'uploaded_by', 'created_at']
    readonly_fields = ['size', 'created_at']


@admin.register(Quote)
class QuoteAdmin(admin.ModelAdmin):
    list_display = ['quote_number', 'company', 'contact', 'status', 'total_value', 'currency', 'valid_until', 'is_expired', 'created_at']
    list_filter = ['status', 'currency', 'created_at', 'valid_until']
    list_select_related = ['company', 'contact', 'opportunity', 'created_by']
    search_fields = ['quote_number', 'company__name', 'contact__name', 'notes']
    readonly_fields = ['quote_number', 'created_by', 'is_expired', 'days_until_expiry', 'created_at', 'updated_at']
    date_hierarchy = 'created_at'
    inlines = [QuoteLineItemInline, QuoteAttachmentInline, QuoteActivityInline]

    fieldsets = (
        ('Quote Information', {
            'fields': ('quote_number', 'company', 'contact', 'opportunity', 'status')
        }),
        ('Validity Period', {
            'fields': ('valid_from', 'valid_until', 'is_expired', 'days_until_expiry')
        }),
        ('Financial Details', {
            'fields': ('subtotal', 'tax_amount', 'discount_amount', 'total_value', 'currency')
        }),
        ('Additional Information', {
            'fields': ('terms_conditions', 'notes'),
            'classes': ('collapse',)
        }),
        ('Status Tracking', {
            'fields': ('sent_date', 'accepted_date', 'rejected_date', 'expired_date'),
            'classes': ('collapse',)
        }),
        ('Metadata', {
            'fields': ('created_by', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    actions = ['mark_as_sent', 'mark_as_accepted', 'mark_as_rejected', 'export_to_excel']

    def save_model(self, request, obj, form, change):
        """Set created_by on new quotes"""
        if not change:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)

    def mark_as_sent(self, request, queryset):
        """Mark selected quotes as sent"""
        count = 0
        for quote in queryset:
            if quote.status == 'Draft':
                quote.status = 'Sent'
                quote.sent_date = timezone.now().date()
                quote.save()

                # Create activity
                QuoteActivity.objects.create(
                    quote=quote,
                    user=request.user,
                    activity_type='Sent',
                    description=f'Quote marked as sent by {request.user.get_full_name()}'
                )
                count += 1

        self.message_user(request, f'{count} quote(s) marked as sent')
    mark_as_sent.short_description = 'Mark selected quotes as sent'

    def mark_as_accepted(self, request, queryset):
        """Mark selected quotes as accepted"""
        count = 0
        for quote in queryset:
            if quote.status in ['Draft', 'Sent']:
                quote.status = 'Accepted'
                quote.accepted_date = timezone.now().date()
                quote.save()

                # Create activity
                QuoteActivity.objects.create(
                    quote=quote,
                    user=request.user,
                    activity_type='Accepted',
                    description=f'Quote marked as accepted by {request.user.get_full_name()}'
                )
                count += 1

        self.message_user(request, f'{count} quote(s) marked as accepted')
    mark_as_accepted.short_description = 'Mark selected quotes as accepted'

    def mark_as_rejected(self, request, queryset):
        """Mark selected quotes as rejected"""
        count = 0
        for quote in queryset:
            if quote.status in ['Draft', 'Sent']:
                quote.status = 'Rejected'
                quote.rejected_date = timezone.now().date()
                quote.save()

                # Create activity
                QuoteActivity.objects.create(
                    quote=quote,
                    user=request.user,
                    activity_type='Rejected',
                    description=f'Quote marked as rejected by {request.user.get_full_name()}'
                )
                count += 1

        self.message_user(request, f'{count} quote(s) marked as rejected')
    mark_as_rejected.short_description = 'Mark selected quotes as rejected'

    def export_to_excel(self, request, queryset):
        """Export selected quotes to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Quotes'

        # Define headers
        headers = [
            'Quote Number', 'Company', 'Contact', 'Status', 'Valid From', 'Valid Until',
            'Subtotal', 'Tax', 'Discount', 'Total', 'Currency', 'Created Date'
        ]

        # Style headers
        header_fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data
        for row_num, quote in enumerate(queryset.select_related('company', 'contact'), 2):
            ws.cell(row=row_num, column=1, value=quote.quote_number)
            ws.cell(row=row_num, column=2, value=quote.company.name)
            ws.cell(row=row_num, column=3, value=quote.contact.name if quote.contact else '')
            ws.cell(row=row_num, column=4, value=quote.status)
            ws.cell(row=row_num, column=5, value=quote.valid_from.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=6, value=quote.valid_until.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=7, value=float(quote.subtotal))
            ws.cell(row=row_num, column=8, value=float(quote.tax_amount))
            ws.cell(row=row_num, column=9, value=float(quote.discount_amount))
            ws.cell(row=row_num, column=10, value=float(quote.total_value))
            ws.cell(row=row_num, column=11, value=quote.currency)
            ws.cell(row=row_num, column=12, value=quote.created_at.strftime('%Y-%m-%d %H:%M'))

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=quotes_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        return response
    export_to_excel.short_description = 'Export selected quotes to Excel'


@admin.register(QuoteLineItem)
class QuoteLineItemAdmin(admin.ModelAdmin):
    list_display = ['quote', 'product_service', 'quantity', 'unit_price', 'discount_percentage', 'line_total']
    list_filter = ['created_at']
    list_select_related = ['quote']
    search_fields = ['product_service', 'description', 'quote__quote_number']
    readonly_fields = ['line_total', 'created_at', 'updated_at']


@admin.register(QuoteAttachment)
class QuoteAttachmentAdmin(admin.ModelAdmin):
    list_display = ['name', 'quote', 'file', 'size', 'uploaded_by', 'created_at']
    list_filter = ['created_at']
    list_select_related = ['quote', 'uploaded_by']
    search_fields = ['name', 'quote__quote_number']
    readonly_fields = ['size', 'created_at', 'updated_at']


@admin.register(QuoteActivity)
class QuoteActivityAdmin(admin.ModelAdmin):
    list_display = ['quote', 'activity_type', 'user', 'description', 'created_at']
    list_filter = ['activity_type', 'created_at']
    list_select_related = ['quote', 'user']
    search_fields = ['quote__quote_number', 'description']
    readonly_fields = ['created_at', 'updated_at']


# ============================================================================
# EMAIL SEQUENCE ADMIN (Drip Campaign System)
# ============================================================================

# INLINE 1: SequenceStepInline (used in EmailSequenceAdmin)
class SequenceStepInline(admin.TabularInline):
    model = SequenceStep
    extra = 1  # Show 1 empty form for adding new step
    fields = ['step_number', 'name', 'email_template', 'delay_days', 'is_active']
    ordering = ['step_number']

    verbose_name = 'Step'
    verbose_name_plural = 'Sequence Steps'


# ADMIN 1: EmailSequenceAdmin
@admin.register(EmailSequence)
class EmailSequenceAdmin(admin.ModelAdmin):
    list_display = ['name', 'status', 'total_steps_display', 'active_enrollments_display', 'created_by', 'created_at']
    list_filter = ['status', 'created_at']
    search_fields = ['name', 'description']
    readonly_fields = ['created_at', 'updated_at', 'created_by']

    fieldsets = [
        ('Sequence Details', {
            'fields': ['name', 'description', 'status']
        }),
        ('Metadata', {
            'fields': ['created_by', 'created_at', 'updated_at'],
            'classes': ['collapse']
        })
    ]

    inlines = [SequenceStepInline]

    # Custom methods for list display
    def total_steps_display(self, obj):
        """Display total number of steps in sequence"""
        return obj.get_total_steps()
    total_steps_display.short_description = 'Total Steps'

    def active_enrollments_display(self, obj):
        """Display count of active enrollments"""
        return obj.get_active_enrollments()
    active_enrollments_display.short_description = 'Active Enrollments'

    # Auto-populate created_by on save
    def save_model(self, request, obj, form, change):
        if not change:  # Only on create
            obj.created_by = request.user
        super().save_model(request, obj, form, change)

    # Bulk actions
    actions = ['activate_sequences', 'pause_sequences', 'archive_sequences']

    def activate_sequences(self, request, queryset):
        """Bulk activate selected sequences"""
        count = queryset.update(status='active')
        self.message_user(request, f'{count} sequence(s) activated', messages.SUCCESS)
    activate_sequences.short_description = 'Activate selected sequences'

    def pause_sequences(self, request, queryset):
        """Bulk pause selected sequences"""
        count = queryset.update(status='paused')
        self.message_user(request, f'{count} sequence(s) paused', messages.SUCCESS)
    pause_sequences.short_description = 'Pause selected sequences'

    def archive_sequences(self, request, queryset):
        """Bulk archive selected sequences"""
        count = queryset.update(status='archived')
        self.message_user(request, f'{count} sequence(s) archived', messages.SUCCESS)
    archive_sequences.short_description = 'Archive selected sequences'


# ADMIN 2: SequenceStepAdmin (Optional Standalone)
@admin.register(SequenceStep)
class SequenceStepAdmin(admin.ModelAdmin):
    list_display = ['__str__', 'sequence', 'step_number', 'email_template', 'delay_days', 'is_active']
    list_filter = ['sequence', 'is_active', 'created_at']
    list_select_related = ['sequence', 'email_template']
    search_fields = ['name', 'sequence__name']
    readonly_fields = ['created_at', 'updated_at']

    fieldsets = [
        ('Step Configuration', {
            'fields': ['sequence', 'step_number', 'name', 'email_template', 'delay_days', 'is_active']
        }),
        ('Metadata', {
            'fields': ['created_at', 'updated_at'],
            'classes': ['collapse']
        })
    ]

    ordering = ['sequence', 'step_number']


# INLINE 2: SequenceStepExecutionInline (used in SequenceEnrollmentAdmin)
class SequenceStepExecutionInline(admin.TabularInline):
    model = SequenceStepExecution
    extra = 0  # Don't show empty forms (executions are auto-created)
    can_delete = False  # Don't allow deletion (audit log)
    fields = ['step', 'scheduled_for', 'sent_at', 'status', 'error_message']
    readonly_fields = ['step', 'scheduled_for', 'sent_at', 'status', 'error_message']
    ordering = ['scheduled_for']

    verbose_name = 'Step Execution'
    verbose_name_plural = 'Execution Log'

    # Make it read-only
    def has_add_permission(self, request, obj=None):
        return False


# ADMIN 3: SequenceEnrollmentAdmin
@admin.register(SequenceEnrollment)
class SequenceEnrollmentAdmin(admin.ModelAdmin):
    list_display = ['contact', 'sequence', 'company', 'status', 'progress_display', 'enrolled_at', 'started_at']
    list_filter = ['status', 'sequence', 'enrolled_at']
    list_select_related = ['contact', 'sequence', 'company']
    search_fields = ['contact__first_name', 'contact__last_name', 'contact__email', 'company__name', 'sequence__name']
    readonly_fields = ['enrolled_at', 'started_at', 'completed_at', 'created_at', 'updated_at']

    fieldsets = [
        ('Enrollment Details', {
            'fields': ['sequence', 'contact', 'company', 'status', 'current_step_number']
        }),
        ('Timeline', {
            'fields': ['enrolled_at', 'started_at', 'completed_at']
        }),
        ('Notes', {
            'fields': ['notes'],
            'classes': ['collapse']
        }),
        ('Metadata', {
            'fields': ['created_at', 'updated_at'],
            'classes': ['collapse']
        })
    ]

    inlines = [SequenceStepExecutionInline]

    # Custom methods
    def progress_display(self, obj):
        """Display enrollment progress (e.g., '2/5 (40.0%)')"""
        return obj.get_progress()
    progress_display.short_description = 'Progress'

    # Bulk actions
    actions = ['pause_enrollments', 'resume_enrollments', 'complete_enrollments']

    def pause_enrollments(self, request, queryset):
        """Bulk pause active enrollments"""
        count = queryset.filter(status='active').update(status='paused')
        self.message_user(request, f'{count} enrollment(s) paused', messages.SUCCESS)
    pause_enrollments.short_description = 'Pause selected enrollments'

    def resume_enrollments(self, request, queryset):
        """Bulk resume paused enrollments"""
        count = queryset.filter(status='paused').update(status='active')
        self.message_user(request, f'{count} enrollment(s) resumed', messages.SUCCESS)
    resume_enrollments.short_description = 'Resume selected enrollments'

    def complete_enrollments(self, request, queryset):
        """Bulk mark enrollments as completed"""
        count = queryset.filter(status='active').update(status='completed', completed_at=timezone.now())
        self.message_user(request, f'{count} enrollment(s) marked complete', messages.SUCCESS)
    complete_enrollments.short_description = 'Mark as completed'


# ADMIN 4: SequenceStepExecutionAdmin
@admin.register(SequenceStepExecution)
class SequenceStepExecutionAdmin(admin.ModelAdmin):
    list_display = ['__str__', 'enrollment_contact', 'enrollment_sequence', 'scheduled_for', 'sent_at', 'status', 'attempt_count']
    list_filter = ['status', 'scheduled_for', 'sent_at']
    list_select_related = ['enrollment__contact', 'enrollment__sequence', 'step', 'email_log']
    search_fields = ['enrollment__contact__email', 'enrollment__sequence__name', 'step__name']
    readonly_fields = ['enrollment', 'step', 'scheduled_for', 'sent_at', 'email_log', 'status', 'error_message', 'attempt_count', 'created_at', 'updated_at']

    fieldsets = [
        ('Execution Details', {
            'fields': ['enrollment', 'step', 'status']
        }),
        ('Scheduling', {
            'fields': ['scheduled_for', 'sent_at', 'attempt_count']
        }),
        ('Results', {
            'fields': ['email_log', 'error_message']
        }),
        ('Metadata', {
            'fields': ['created_at', 'updated_at'],
            'classes': ['collapse']
        })
    ]

    ordering = ['-scheduled_for']

    # Custom methods for better display
    def enrollment_contact(self, obj):
        """Display contact name from enrollment"""
        return obj.enrollment.contact
    enrollment_contact.short_description = 'Contact'

    def enrollment_sequence(self, obj):
        """Display sequence name from enrollment"""
        return obj.enrollment.sequence
    enrollment_sequence.short_description = 'Sequence'

    # Read-only admin (audit log)
    def has_add_permission(self, request):
        """Prevent manual creation of executions"""
        return False

    def has_delete_permission(self, request, obj=None):
        """Prevent deletion of audit records"""
        return False


# Customer Segmentation Admin
@admin.register(CustomerSegment)
class CustomerSegmentAdmin(admin.ModelAdmin):
    list_display = ['name', 'segment_type', 'status', 'member_count', 'created_by', 'created_at']
    list_filter = ['segment_type', 'status', 'created_at']
    search_fields = ['name', 'description', 'tags']
    readonly_fields = ['member_count', 'last_calculated_at', 'last_used_at', 'times_used', 'created_at', 'updated_at']

    fieldsets = (
        ('Basic Information', {
            'fields': ('name', 'description', 'segment_type', 'status', 'tags')
        }),
        ('Filter Criteria (Dynamic Segments)', {
            'fields': ('filter_criteria',),
            'classes': ('collapse',),
        }),
        ('Static Contacts (Static Segments)', {
            'fields': ('static_contacts', 'static_companies'),
            'classes': ('collapse',),
        }),
        ('Statistics', {
            'fields': ('member_count', 'last_calculated_at', 'last_used_at', 'times_used')
        }),
        ('Metadata', {
            'fields': ('created_by', 'created_at', 'updated_at')
        }),
    )

    filter_horizontal = ('static_contacts', 'static_companies')

    def save_model(self, request, obj, form, change):
        if not change:  # New object
            obj.created_by = request.user
        super().save_model(request, obj, form, change)

        # Calculate member count after save
        if obj.segment_type == 'dynamic':
            obj.update_member_count()


# Support Ticket System Admin
class TicketCommentInline(admin.TabularInline):
    """Inline admin for ticket comments"""
    model = TicketComment
    extra = 1
    fields = ['author', 'text', 'is_internal', 'created_at']
    readonly_fields = ['created_at']

    def get_queryset(self, request):
        """Optimize inline queries"""
        return super().get_queryset(request).select_related('author', 'ticket')


class TicketAttachmentInline(admin.TabularInline):
    """Inline admin for ticket attachments"""
    model = TicketAttachment
    extra = 1
    fields = ['file', 'name', 'size_display', 'uploaded_by', 'created_at']
    readonly_fields = ['size_display', 'created_at']

    def size_display(self, obj):
        """Display file size in human-readable format"""
        if obj.size:
            if obj.size < 1024:
                return f"{obj.size} bytes"
            elif obj.size < 1024 * 1024:
                return f"{obj.size / 1024:.2f} KB"
            else:
                return f"{obj.size / (1024 * 1024):.2f} MB"
        return "Unknown"
    size_display.short_description = "File Size"

    def get_queryset(self, request):
        """Optimize inline queries"""
        return super().get_queryset(request).select_related('uploaded_by', 'ticket')


@admin.register(Ticket)
class TicketAdmin(admin.ModelAdmin):
    """Admin interface for support tickets"""
    list_display = [
        'ticket_number',
        'subject',
        'company',
        'status_badge',
        'priority_badge',
        'category',
        'assigned_to',
        'created_at',
        'is_overdue_badge'
    ]
    list_filter = [
        'status',
        'priority',
        'category',
        'assigned_team',
        'created_at',
        'assigned_to'
    ]
    search_fields = [
        'ticket_number',
        'subject',
        'description',
        'company__name',
        'contact__name',
        'contact__email'
    ]
    readonly_fields = [
        'ticket_number',
        'created_by',
        'created_at',
        'updated_at',
        'first_response_time_display',
        'resolution_time_display',
        'is_overdue_badge'
    ]

    fieldsets = (
        ('Ticket Information', {
            'fields': (
                'ticket_number',
                'subject',
                'description',
                'company',
                'contact'
            )
        }),
        ('Classification', {
            'fields': (
                'status',
                'priority',
                'category',
                'tags'
            )
        }),
        ('Assignment', {
            'fields': (
                'assigned_to',
                'assigned_team',
                'created_by'
            )
        }),
        ('Time Tracking', {
            'fields': (
                'due_date',
                'first_response_at',
                'first_response_time_display',
                'resolved_at',
                'resolution_time_display',
                'closed_at',
                'is_overdue_badge'
            )
        }),
        ('Metadata', {
            'fields': (
                'created_at',
                'updated_at'
            ),
            'classes': ('collapse',)
        }),
    )

    inlines = [TicketCommentInline, TicketAttachmentInline]

    def get_queryset(self, request):
        """Optimize queryset with select_related"""
        return super().get_queryset(request).select_related(
            'company',
            'contact',
            'assigned_to',
            'created_by'
        )

    def status_badge(self, obj):
        """Display status with color badge"""
        colors = {
            'new': '#e74c3c',  # Red
            'assigned': '#f39c12',  # Orange
            'in_progress': '#3498db',  # Blue
            'pending': '#9b59b6',  # Purple
            'resolved': '#2ecc71',  # Green
            'closed': '#95a5a6',  # Gray
        }
        color = colors.get(obj.status, '#95a5a6')
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; border-radius: 3px; font-weight: bold;">{}</span>',
            color,
            obj.get_status_display()
        )
    status_badge.short_description = 'Status'

    def priority_badge(self, obj):
        """Display priority with color badge"""
        colors = {
            'low': '#95a5a6',  # Gray
            'medium': '#3498db',  # Blue
            'high': '#f39c12',  # Orange
            'urgent': '#e74c3c',  # Red
        }
        color = colors.get(obj.priority, '#95a5a6')
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; border-radius: 3px; font-weight: bold;">{}</span>',
            color,
            obj.get_priority_display()
        )
    priority_badge.short_description = 'Priority'

    def is_overdue_badge(self, obj):
        """Display overdue status"""
        if obj.is_overdue:
            return format_html(
                '<span style="background-color: #e74c3c; color: white; padding: 3px 8px; border-radius: 3px; font-weight: bold;">OVERDUE</span>'
            )
        return format_html(
            '<span style="color: #2ecc71; font-weight: bold;">On Track</span>'
        )
    is_overdue_badge.short_description = 'Status'

    def first_response_time_display(self, obj):
        """Display first response time in hours"""
        hours = obj.first_response_time_hours
        if hours is not None:
            return f"{hours} hours"
        return "No response yet"
    first_response_time_display.short_description = 'First Response Time'

    def resolution_time_display(self, obj):
        """Display resolution time in hours"""
        hours = obj.resolution_time_hours
        if hours is not None:
            return f"{hours} hours"
        return "Not resolved"
    resolution_time_display.short_description = 'Resolution Time'

    def save_model(self, request, obj, form, change):
        """Set created_by on new tickets"""
        if not change:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(TicketComment)
class TicketCommentAdmin(admin.ModelAdmin):
    """Admin interface for ticket comments"""
    list_display = ['ticket', 'author', 'comment_preview', 'is_internal', 'created_at']
    list_filter = ['is_internal', 'created_at']
    search_fields = ['ticket__ticket_number', 'ticket__subject', 'text', 'author__username']
    readonly_fields = ['created_at', 'updated_at']

    def get_queryset(self, request):
        """Optimize queryset"""
        return super().get_queryset(request).select_related('ticket', 'author')

    def comment_preview(self, obj):
        """Show preview of comment text"""
        return obj.text[:100] + '...' if len(obj.text) > 100 else obj.text
    comment_preview.short_description = 'Comment'


@admin.register(TicketAttachment)
class TicketAttachmentAdmin(admin.ModelAdmin):
    """Admin interface for ticket attachments"""
    list_display = ['name', 'ticket', 'size_display', 'uploaded_by', 'created_at']
    list_filter = ['created_at']
    search_fields = ['name', 'ticket__ticket_number', 'ticket__subject']
    readonly_fields = ['size', 'created_at']

    def get_queryset(self, request):
        """Optimize queryset"""
        return super().get_queryset(request).select_related('ticket', 'uploaded_by')

    def size_display(self, obj):
        """Display file size in human-readable format"""
        if obj.size:
            if obj.size < 1024:
                return f"{obj.size} bytes"
            elif obj.size < 1024 * 1024:
                return f"{obj.size / 1024:.2f} KB"
            else:
                return f"{obj.size / (1024 * 1024):.2f} MB"
        return "Unknown"
    size_display.short_description = 'File Size'
# =============================================================================
# KNOWLEDGE BASE SYSTEM ADMIN INTERFACES
# =============================================================================


# -----------------------------------------------------------------------------
# KBCategory Admin with Hierarchical Categories Support
# -----------------------------------------------------------------------------

class ChildCategoryInline(admin.TabularInline):
    """Inline editor for child categories"""
    model = KBCategory
    fk_name = 'parent'
    extra = 0
    fields = ['name', 'slug', 'display_order', 'is_active', 'icon']
    prepopulated_fields = {'slug': ('name',)}


@admin.register(KBCategory)
class KBCategoryAdmin(admin.ModelAdmin):
    """Admin interface for KB categories with hierarchical support"""
    list_display = ['name', 'parent', 'article_count_display', 'display_order', 'is_active', 'icon']
    list_filter = ['is_active', 'parent']
    search_fields = ['name', 'description', 'slug']
    prepopulated_fields = {'slug': ('name',)}
    readonly_fields = ['created_at', 'updated_at']
    list_per_page = 25
    inlines = [ChildCategoryInline]
    actions = ['activate_categories', 'deactivate_categories']

    fieldsets = (
        ('Basic Information', {
            'fields': ('name', 'slug', 'description', 'parent')
        }),
        ('Display Settings', {
            'fields': ('icon', 'display_order', 'is_active')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def article_count_display(self, obj):
        """Display count of published articles in this category"""
        count = obj.article_count
        if count > 0:
            return format_html(
                '<span style="background-color: #4CAF50; color: white; padding: 3px 8px; '
                'border-radius: 3px; font-weight: bold;">{}</span>',
                count
            )
        return format_html('<span style="color: #999;">0</span>')
    article_count_display.short_description = 'Published Articles'

    def activate_categories(self, request, queryset):
        """Bulk activate selected categories"""
        count = queryset.update(is_active=True)
        self.message_user(request, f'Successfully activated {count} categories')
    activate_categories.short_description = 'Activate selected categories'

    def deactivate_categories(self, request, queryset):
        """Bulk deactivate selected categories"""
        count = queryset.update(is_active=False)
        self.message_user(request, f'Successfully deactivated {count} categories')
    deactivate_categories.short_description = 'Deactivate selected categories'


# -----------------------------------------------------------------------------
# KBTag Admin with Color Preview
# -----------------------------------------------------------------------------

@admin.register(KBTag)
class KBTagAdmin(admin.ModelAdmin):
    """Admin interface for KB tags with color preview"""
    list_display = ['name', 'color_preview', 'article_count_display', 'slug']
    search_fields = ['name', 'slug']
    prepopulated_fields = {'slug': ('name',)}
    readonly_fields = ['created_at', 'updated_at']
    list_per_page = 50

    fieldsets = (
        ('Tag Information', {
            'fields': ('name', 'slug', 'color')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def color_preview(self, obj):
        """Display color with visual preview"""
        return format_html(
            '<span style="display: inline-block; width: 80px; padding: 5px 10px; '
            'background-color: {}; color: white; border-radius: 3px; text-align: center;">{}</span>',
            obj.color, obj.color
        )
    color_preview.short_description = 'Color'

    def article_count_display(self, obj):
        """Display count of published articles with this tag"""
        count = obj.article_count
        if count > 0:
            return format_html(
                '<span style="background-color: #2196F3; color: white; padding: 3px 8px; '
                'border-radius: 3px; font-weight: bold;">{}</span>',
                count
            )
        return format_html('<span style="color: #999;">0</span>')
    article_count_display.short_description = 'Articles'


# -----------------------------------------------------------------------------
# KBArticle Admin - Main Knowledge Base Interface
# -----------------------------------------------------------------------------

class KBArticleAttachmentInline(admin.TabularInline):
    """Inline editor for article attachments"""
    model = KBArticleAttachment
    extra = 0
    fields = ['file', 'filename', 'file_size_display', 'uploaded_by']
    readonly_fields = ['filename', 'file_size_display', 'uploaded_by']

    def file_size_display(self, obj):
        """Display file size in human-readable format"""
        if obj.file_size:
            if obj.file_size < 1024:
                return f"{obj.file_size} bytes"
            elif obj.file_size < 1024 * 1024:
                return f"{obj.file_size / 1024:.2f} KB"
            else:
                return f"{obj.file_size / (1024 * 1024):.2f} MB"
        return "Unknown"
    file_size_display.short_description = 'File Size'


class KBArticleRelationInline(admin.TabularInline):
    """Inline editor for related articles"""
    model = KBArticleRelation
    fk_name = 'from_article'
    extra = 0
    fields = ['to_article', 'relation_type', 'display_order']
    autocomplete_fields = ['to_article']


@admin.register(KBArticle)
class KBArticleAdmin(admin.ModelAdmin):
    """Comprehensive admin interface for KB articles"""
    list_display = [
        'article_number', 'title', 'category', 'status_badge', 'visibility_badge',
        'author', 'view_count_display', 'helpfulness_display', 'published_at', 'featured_badge'
    ]
    list_filter = ['status', 'visibility', 'category', 'featured', 'tags', 'created_at']
    search_fields = ['article_number', 'title', 'content', 'excerpt']
    prepopulated_fields = {'slug': ('title',)}
    # Enable autocomplete for use in relations and ticket linking
    autocomplete_fields = []
    readonly_fields = [
        'article_number', 'view_count', 'helpful_count', 'not_helpful_count',
        'search_vector', 'created_at', 'updated_at', 'preview_link'
    ]
    filter_horizontal = ['tags']
    list_per_page = 25
    date_hierarchy = 'created_at'
    inlines = [KBArticleAttachmentInline, KBArticleRelationInline]
    actions = [
        'publish_articles', 'archive_articles', 'mark_as_featured',
        'remove_from_featured', 'export_articles_csv'
    ]

    fieldsets = (
        ('Content', {
            'fields': ('title', 'slug', 'content', 'excerpt', 'category', 'tags')
        }),
        ('Publishing', {
            'fields': ('status', 'visibility', 'author', 'featured', 'published_at', 'preview_link')
        }),
        ('Metadata', {
            'fields': ('article_number',),
            'classes': ('collapse',)
        }),
        ('Statistics', {
            'fields': ('view_count', 'helpful_count', 'not_helpful_count'),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def get_queryset(self, request):
        """Optimize queryset with related data"""
        return super().get_queryset(request).select_related(
            'category', 'author'
        ).prefetch_related('tags')

    def status_badge(self, obj):
        """Display status with color coding"""
        colors = {
            'draft': '#FF9800',
            'published': '#4CAF50',
            'archived': '#9E9E9E',
        }
        color = colors.get(obj.status, '#999')
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; '
            'border-radius: 3px; font-weight: bold;">{}</span>',
            color, obj.get_status_display()
        )
    status_badge.short_description = 'Status'

    def visibility_badge(self, obj):
        """Display visibility with color coding"""
        colors = {
            'public': '#2196F3',
            'internal': '#9C27B0',
        }
        color = colors.get(obj.visibility, '#999')
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 6px; '
            'border-radius: 3px; font-size: 11px;">{}</span>',
            color, obj.get_visibility_display()
        )
    visibility_badge.short_description = 'Visibility'

    def featured_badge(self, obj):
        """Display featured status"""
        if obj.featured:
            return format_html(
                '<span style="color: #FFA500; font-size: 16px;" title="Featured">★</span>'
            )
        return format_html('<span style="color: #CCC;">☆</span>')
    featured_badge.short_description = 'Featured'

    def view_count_display(self, obj):
        """Display view count with formatting"""
        if obj.view_count > 100:
            return format_html('<strong style="color: #4CAF50;">{}</strong>', obj.view_count)
        elif obj.view_count > 10:
            return format_html('<span style="color: #FF9800;">{}</span>', obj.view_count)
        return str(obj.view_count)
    view_count_display.short_description = 'Views'

    def helpfulness_display(self, obj):
        """Display helpfulness as percentage with color coding"""
        ratio = obj.get_helpfulness_ratio()
        if ratio is None:
            return format_html('<span style="color: #999;">No votes</span>')

        # Color coding: green >70%, yellow 40-70%, red <40%
        if ratio >= 70:
            color = '#4CAF50'
        elif ratio >= 40:
            color = '#FF9800'
        else:
            color = '#F44336'

        total_votes = obj.helpful_count + obj.not_helpful_count
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; '
            'border-radius: 3px; font-weight: bold;">{:.1f}%</span> '
            '<span style="color: #999; font-size: 11px;">({} votes)</span>',
            color, ratio, total_votes
        )
    helpfulness_display.short_description = 'Helpfulness'

    def preview_link(self, obj):
        """Display article preview link if published"""
        if obj.status == 'published':
            # Assuming frontend URL pattern: /kb/articles/{slug}
            url = f"/kb/articles/{obj.slug}"
            return format_html(
                '<a href="{}" target="_blank" style="color: #2196F3; text-decoration: underline;">'
                'View Article →</a>',
                url
            )
        return format_html('<span style="color: #999;">Not published yet</span>')
    preview_link.short_description = 'Preview'

    def publish_articles(self, request, queryset):
        """Publish selected articles"""
        count = queryset.filter(status='draft').update(
            status='published',
            published_at=timezone.now()
        )
        self.message_user(request, f'Successfully published {count} articles')
    publish_articles.short_description = 'Publish selected articles'

    def archive_articles(self, request, queryset):
        """Archive selected articles"""
        count = queryset.update(status='archived')
        self.message_user(request, f'Successfully archived {count} articles')
    archive_articles.short_description = 'Archive selected articles'

    def mark_as_featured(self, request, queryset):
        """Mark selected articles as featured"""
        count = queryset.update(featured=True)
        self.message_user(request, f'Successfully marked {count} articles as featured')
    mark_as_featured.short_description = 'Mark as featured'

    def remove_from_featured(self, request, queryset):
        """Remove selected articles from featured"""
        count = queryset.update(featured=False)
        self.message_user(request, f'Successfully removed {count} articles from featured')
    remove_from_featured.short_description = 'Remove from featured'

    def export_articles_csv(self, request, queryset):
        """Export selected articles to CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="kb_articles_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'Article Number', 'Title', 'Category', 'Status', 'Visibility',
            'Author', 'Views', 'Helpful Votes', 'Not Helpful Votes', 'Helpfulness %',
            'Featured', 'Published Date', 'Created Date'
        ])

        articles = queryset.select_related('category', 'author')
        for article in articles:
            ratio = article.get_helpfulness_ratio()
            writer.writerow([
                article.article_number,
                article.title,
                article.category.name,
                article.get_status_display(),
                article.get_visibility_display(),
                article.author.username if article.author else 'N/A',
                article.view_count,
                article.helpful_count,
                article.not_helpful_count,
                f'{ratio:.1f}%' if ratio is not None else 'N/A',
                'Yes' if article.featured else 'No',
                article.published_at.strftime('%Y-%m-%d %H:%M') if article.published_at else 'Not published',
                article.created_at.strftime('%Y-%m-%d %H:%M')
            ])

        self.message_user(request, f'Successfully exported {queryset.count()} articles to CSV')
        return response
    export_articles_csv.short_description = 'Export to CSV'

    def save_model(self, request, obj, form, change):
        """Auto-populate author from request.user on create"""
        if not change and not obj.author:
            obj.author = request.user
        super().save_model(request, obj, form, change)


# -----------------------------------------------------------------------------
# KBArticleView Admin - Analytics
# -----------------------------------------------------------------------------

@admin.register(KBArticleView)
class KBArticleViewAdmin(admin.ModelAdmin):
    """Admin interface for KB article views (analytics only)"""
    list_display = ['article_number', 'article_title', 'user_display', 'ip_address', 'viewed_at']
    list_filter = ['viewed_at', 'user']
    search_fields = ['article__title', 'article__article_number', 'ip_address']
    readonly_fields = ['article', 'user', 'ip_address', 'session_id', 'viewed_at']
    date_hierarchy = 'viewed_at'
    list_per_page = 50

    fieldsets = (
        ('View Information', {
            'fields': ('article', 'user', 'ip_address', 'session_id', 'viewed_at')
        }),
    )

    def get_queryset(self, request):
        """Optimize queryset"""
        return super().get_queryset(request).select_related('article', 'user')

    def article_number(self, obj):
        """Display article number"""
        return obj.article.article_number
    article_number.short_description = 'Article #'

    def article_title(self, obj):
        """Display article title with link"""
        url = reverse('admin:crm_app_kbarticle_change', args=[obj.article.pk])
        return format_html('<a href="{}">{}</a>', url, obj.article.title[:50])
    article_title.short_description = 'Article'

    def user_display(self, obj):
        """Display user or anonymous"""
        if obj.user:
            return obj.user.username
        return format_html('<span style="color: #999;">Anonymous</span>')
    user_display.short_description = 'User'

    def has_add_permission(self, request):
        """Disable manual adding (analytics only)"""
        return False

    def has_delete_permission(self, request, obj=None):
        """Allow deletion for data cleanup"""
        return True


# -----------------------------------------------------------------------------
# KBArticleRating Admin
# -----------------------------------------------------------------------------

@admin.register(KBArticleRating)
class KBArticleRatingAdmin(admin.ModelAdmin):
    """Admin interface for KB article ratings"""
    list_display = ['article_number', 'article_title', 'user', 'vote_badge', 'created_at']
    list_filter = ['is_helpful', 'created_at']
    search_fields = ['article__title', 'article__article_number', 'user__username']
    readonly_fields = ['article', 'user', 'is_helpful', 'created_at']
    date_hierarchy = 'created_at'
    list_per_page = 50

    fieldsets = (
        ('Rating Information', {
            'fields': ('article', 'user', 'is_helpful', 'created_at')
        }),
    )

    def get_queryset(self, request):
        """Optimize queryset"""
        return super().get_queryset(request).select_related('article', 'user')

    def article_number(self, obj):
        """Display article number"""
        return obj.article.article_number
    article_number.short_description = 'Article #'

    def article_title(self, obj):
        """Display article title with link"""
        url = reverse('admin:crm_app_kbarticle_change', args=[obj.article.pk])
        return format_html('<a href="{}">{}</a>', url, obj.article.title[:50])
    article_title.short_description = 'Article'

    def vote_badge(self, obj):
        """Display vote type with color"""
        if obj.is_helpful:
            return format_html(
                '<span style="background-color: #4CAF50; color: white; padding: 3px 8px; '
                'border-radius: 3px; font-weight: bold;">👍 Helpful</span>'
            )
        return format_html(
            '<span style="background-color: #F44336; color: white; padding: 3px 8px; '
            'border-radius: 3px; font-weight: bold;">👎 Not Helpful</span>'
        )
    vote_badge.short_description = 'Vote'

    def has_add_permission(self, request):
        """Disable manual adding"""
        return False


# -----------------------------------------------------------------------------
# KBArticleRelation Admin
# -----------------------------------------------------------------------------

@admin.register(KBArticleRelation)
class KBArticleRelationAdmin(admin.ModelAdmin):
    """Admin interface for related articles"""
    list_display = ['from_article_number', 'relation_type_badge', 'to_article_number', 'display_order']
    list_filter = ['relation_type']
    search_fields = ['from_article__title', 'to_article__title', 'from_article__article_number', 'to_article__article_number']
    autocomplete_fields = ['from_article', 'to_article']
    list_per_page = 50

    fieldsets = (
        ('Relation Information', {
            'fields': ('from_article', 'to_article', 'relation_type', 'display_order')
        }),
    )

    def get_queryset(self, request):
        """Optimize queryset"""
        return super().get_queryset(request).select_related('from_article', 'to_article')

    def from_article_number(self, obj):
        """Display source article number"""
        return obj.from_article.article_number
    from_article_number.short_description = 'From Article'

    def to_article_number(self, obj):
        """Display target article number"""
        return obj.to_article.article_number
    to_article_number.short_description = 'To Article'

    def relation_type_badge(self, obj):
        """Display relation type with styling"""
        colors = {
            'related': '#2196F3',
            'see_also': '#9C27B0',
            'prerequisite': '#FF9800',
        }
        color = colors.get(obj.relation_type, '#999')
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; '
            'border-radius: 3px; font-size: 11px;">{}</span>',
            color, obj.get_relation_type_display()
        )
    relation_type_badge.short_description = 'Relation Type'


# -----------------------------------------------------------------------------
# KBArticleAttachment Admin
# -----------------------------------------------------------------------------

@admin.register(KBArticleAttachment)
class KBArticleAttachmentAdmin(admin.ModelAdmin):
    """Admin interface for KB article attachments"""
    list_display = ['filename', 'article_number', 'file_size_display', 'file_extension', 'uploaded_by', 'uploaded_at']
    list_filter = ['uploaded_at']
    search_fields = ['filename', 'article__title', 'article__article_number']
    readonly_fields = ['filename', 'file_size', 'uploaded_by', 'uploaded_at']
    date_hierarchy = 'uploaded_at'
    list_per_page = 50

    fieldsets = (
        ('Attachment Information', {
            'fields': ('article', 'file', 'filename', 'file_size')
        }),
        ('Upload Details', {
            'fields': ('uploaded_by', 'uploaded_at')
        }),
    )

    def get_queryset(self, request):
        """Optimize queryset"""
        return super().get_queryset(request).select_related('article', 'uploaded_by')

    def article_number(self, obj):
        """Display article number"""
        return obj.article.article_number
    article_number.short_description = 'Article #'

    def file_size_display(self, obj):
        """Display file size in human-readable format"""
        if obj.file_size:
            if obj.file_size < 1024:
                return f"{obj.file_size} bytes"
            elif obj.file_size < 1024 * 1024:
                return f"{obj.file_size / 1024:.2f} KB"
            elif obj.file_size < 1024 * 1024 * 1024:
                return f"{obj.file_size / (1024 * 1024):.2f} MB"
            else:
                return f"{obj.file_size / (1024 * 1024 * 1024):.2f} GB"
        return "Unknown"
    file_size_display.short_description = 'File Size'

    def file_extension(self, obj):
        """Display file extension with icon-style badge"""
        ext = obj.get_file_extension().upper().replace('.', '')

        # Color coding by file type
        color_map = {
            'PDF': '#F44336',
            'DOC': '#2196F3', 'DOCX': '#2196F3',
            'XLS': '#4CAF50', 'XLSX': '#4CAF50',
            'JPG': '#FF9800', 'JPEG': '#FF9800', 'PNG': '#FF9800', 'GIF': '#FF9800',
            'ZIP': '#9C27B0', 'RAR': '#9C27B0',
        }
        color = color_map.get(ext, '#607D8B')

        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 6px; '
            'border-radius: 3px; font-size: 10px; font-weight: bold;">{}</span>',
            color, ext if ext else 'FILE'
        )
    file_extension.short_description = 'Type'

    def save_model(self, request, obj, form, change):
        """Auto-populate uploaded_by from request.user on create"""
        if not change and not obj.uploaded_by:
            obj.uploaded_by = request.user
        super().save_model(request, obj, form, change)


# -----------------------------------------------------------------------------
# TicketKBArticle Admin
# -----------------------------------------------------------------------------

@admin.register(TicketKBArticle)
class TicketKBArticleAdmin(admin.ModelAdmin):
    """Admin interface for ticket-article links"""
    list_display = ['ticket_number', 'article_number', 'linked_by', 'helpful_badge', 'linked_at']
    list_filter = ['is_helpful', 'linked_at']
    search_fields = ['ticket__ticket_number', 'article__article_number', 'article__title']
    autocomplete_fields = ['ticket', 'article']
    readonly_fields = ['linked_by', 'linked_at']
    date_hierarchy = 'linked_at'
    list_per_page = 50

    fieldsets = (
        ('Link Information', {
            'fields': ('ticket', 'article', 'linked_by', 'linked_at')
        }),
        ('Feedback', {
            'fields': ('is_helpful',),
            'description': 'Was this article helpful for resolving the ticket?'
        }),
    )

    def get_queryset(self, request):
        """Optimize queryset"""
        return super().get_queryset(request).select_related('ticket', 'article', 'linked_by')

    def ticket_number(self, obj):
        """Display ticket number with link"""
        url = reverse('admin:crm_app_ticket_change', args=[obj.ticket.pk])
        return format_html('<a href="{}">{}</a>', url, obj.ticket.ticket_number)
    ticket_number.short_description = 'Ticket'

    def article_number(self, obj):
        """Display article number with link"""
        url = reverse('admin:crm_app_kbarticle_change', args=[obj.article.pk])
        return format_html('<a href="{}">{}</a>', url, obj.article.article_number)
    article_number.short_description = 'KB Article'

    def helpful_badge(self, obj):
        """Display helpfulness status"""
        if obj.is_helpful is None:
            return format_html('<span style="color: #999;">Not rated</span>')
        elif obj.is_helpful:
            return format_html(
                '<span style="background-color: #4CAF50; color: white; padding: 3px 8px; '
                'border-radius: 3px; font-weight: bold;">✓ Helpful</span>'
            )
        else:
            return format_html(
                '<span style="background-color: #F44336; color: white; padding: 3px 8px; '
                'border-radius: 3px; font-weight: bold;">✗ Not Helpful</span>'
            )
    helpful_badge.short_description = 'Helpful?'

    def save_model(self, request, obj, form, change):
        """Auto-populate linked_by from request.user on create"""
        if not change and not obj.linked_by:
            obj.linked_by = request.user
        super().save_model(request, obj, form, change)


# -----------------------------------------------------------------------------
# Equipment Management Admin
# -----------------------------------------------------------------------------

# Equipment admin classes removed - replaced by simpler Device model


# -----------------------------------------------------------------------------
# Static Documents Admin (Terms & Conditions, etc.)
# -----------------------------------------------------------------------------

@admin.register(StaticDocument)
class StaticDocumentAdmin(admin.ModelAdmin):
    """Admin interface for managing static legal documents"""
    list_display = ['name', 'document_type_badge', 'version', 'effective_date', 'is_active_badge', 'file_link', 'updated_at']
    list_filter = ['document_type', 'is_active', 'effective_date']
    search_fields = ['name', 'description', 'version']
    readonly_fields = ['created_at', 'updated_at', 'file_size_display']
    list_per_page = 25
    date_hierarchy = 'effective_date'

    fieldsets = (
        ('Document Information', {
            'fields': ('name', 'document_type', 'version', 'effective_date', 'is_active')
        }),
        ('File', {
            'fields': ('file', 'file_size_display'),
            'description': 'Upload PDF document. Recommended: Use clear naming and keep file sizes under 5MB.'
        }),
        ('Description', {
            'fields': ('description',),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def document_type_badge(self, obj):
        """Display document type with colored badge"""
        color_map = {
            'standard_terms_th': '#FFA500',  # Orange for Thailand/HK
            'standard_terms_intl': '#2196F3',  # Blue for International
        }
        label_map = {
            'standard_terms_th': 'TH/HK',
            'standard_terms_intl': 'INTL',
        }
        color = color_map.get(obj.document_type, '#999')
        label = label_map.get(obj.document_type, obj.get_document_type_display())

        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; '
            'border-radius: 3px; font-size: 11px; font-weight: bold;">{}</span><br>'
            '<small style="color: #666;">{}</small>',
            color, label, obj.get_document_type_display()
        )
    document_type_badge.short_description = "Type"

    def is_active_badge(self, obj):
        """Display active status with colored badge"""
        if obj.is_active:
            return format_html(
                '<span style="background-color: #4CAF50; color: white; padding: 3px 8px; '
                'border-radius: 3px; font-size: 11px; font-weight: bold;">ACTIVE</span>'
            )
        return format_html(
            '<span style="background-color: #999; color: white; padding: 3px 8px; '
            'border-radius: 3px; font-size: 11px;">INACTIVE</span>'
        )
    is_active_badge.short_description = "Status"

    def file_link(self, obj):
        """Display link to download the file"""
        if obj.file:
            return format_html(
                '<a href="{}" target="_blank" style="color: #2196F3; text-decoration: none;">'
                '📄 View PDF</a>',
                obj.file.url
            )
        return format_html('<span style="color: #999;">No file</span>')
    file_link.short_description = "File"

    def file_size_display(self, obj):
        """Display file size in human-readable format"""
        if obj.file:
            try:
                size = obj.file.size
                if size < 1024:
                    return f"{size} bytes"
                elif size < 1024 * 1024:
                    return f"{size / 1024:.2f} KB"
                else:
                    return f"{size / (1024 * 1024):.2f} MB"
            except Exception:
                return "Unknown size"
        return "No file"
    file_size_display.short_description = "File Size"

    def get_queryset(self, request):
        """Optimize queryset"""
        return super().get_queryset(request).order_by('-is_active', '-effective_date')


# ============================================================================
# CONTRACT CONTENT MANAGEMENT ADMIN INTERFACES
# ============================================================================


@admin.register(ContractTemplate)
class ContractTemplateAdmin(admin.ModelAdmin):
    """Admin interface for contract templates with variable substitution"""
    list_display = ['name', 'template_type', 'version', 'is_default', 'is_active', 'updated_at']
    list_filter = ['template_type', 'is_default', 'is_active']
    search_fields = ['name', 'content']
    ordering = ['template_type', 'name']
    readonly_fields = ['created_at', 'updated_at']

    fieldsets = (
        (None, {
            'fields': ('name', 'template_type', 'version')
        }),
        ('Content', {
            'fields': ('content',),
            'description': 'Use {{variables}} for substitution: {{company_name}}, {{contract_number}}, {{start_date}}, {{end_date}}, {{value}}, {{currency}}, {{billing_frequency}}, {{payment_terms}}'
        }),
        ('Settings', {
            'fields': ('is_default', 'is_active')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )


@admin.register(ServicePackageItem)
class ServicePackageItemAdmin(admin.ModelAdmin):
    """Admin interface for service package items"""
    list_display = ['name', 'display_order', 'is_standard', 'created_at']
    list_filter = ['is_standard']
    list_editable = ['display_order', 'is_standard']
    search_fields = ['name', 'description']
    ordering = ['display_order', 'name']


@admin.register(CorporatePdfTemplate)
class CorporatePdfTemplateAdmin(admin.ModelAdmin):
    """Admin interface for corporate PDF templates"""
    list_display = ['name', 'company', 'template_format', 'include_attachment_a', 'include_exhibit_d', 'updated_at']
    list_filter = ['template_format']
    search_fields = ['name', 'company__name']
    autocomplete_fields = ['company']
    readonly_fields = ['created_at', 'updated_at']

    fieldsets = (
        (None, {
            'fields': ('name', 'company', 'template_format')
        }),
        ('Document Options', {
            'fields': ('include_attachment_a', 'include_exhibit_d')
        }),
        ('Custom Content', {
            'fields': ('header_text', 'legal_terms', 'warranty_text'),
            'classes': ('collapse',)
        }),
        ('Branding', {
            'fields': ('use_corporate_branding', 'corporate_logo'),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )


@admin.register(ContractDocument)
class ContractDocumentAdmin(admin.ModelAdmin):
    """Admin interface for contract documents"""
    list_display = ['title', 'contract', 'document_type', 'is_official', 'is_signed', 'uploaded_at']
    list_filter = ['document_type', 'is_official', 'is_signed']
    search_fields = ['title', 'contract__contract_number', 'notes']
    autocomplete_fields = ['contract']
    readonly_fields = ['uploaded_at', 'uploaded_by']
    date_hierarchy = 'uploaded_at'

    fieldsets = (
        (None, {
            'fields': ('contract', 'document_type', 'title')
        }),
        ('File', {
            'fields': ('file',)
        }),
        ('Status', {
            'fields': ('is_official', 'is_signed', 'signed_date')
        }),
        ('Additional Information', {
            'fields': ('notes', 'uploaded_at', 'uploaded_by'),
            'classes': ('collapse',)
        }),
    )
