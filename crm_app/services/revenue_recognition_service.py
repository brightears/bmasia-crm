"""
Revenue Recognition Service for BMAsia CRM

Implements accrual-based revenue recognition matching Pom's spreadsheet format.
Revenue is recognized based on actual calendar days within the service period.

Formula (verified against Pom's Excel):
    daily_rate = invoice_amount / total_service_days
    quarterly_income = daily_rate × service_days_in_quarter
    balance = invoice_amount - cumulative_recognized_income

Service days = actual calendar days (inclusive: end - start + 1).
Revenue must NOT be recognized before invoice date.

Usage:
    from crm_app.services.revenue_recognition_service import RevenueRecognitionService

    service = RevenueRecognitionService()
    service.generate_entries_for_year(2025, 'bmasia_hk')
    summary = service.get_quarterly_summary(2025, 'bmasia_hk')
"""

import logging
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from typing import Dict, List, Optional, Tuple
from django.db import transaction
from django.db.models import Sum, Q, Count

logger = logging.getLogger(__name__)

# Product classification keywords (reuses pattern from quickbooks_export_service.py)
PRODUCT_KEYWORDS = {
    'SYB': ['soundtrack', 'syb', 'stb', 'soundtrack your brand'],
    'LIM': ['lim', 'beatbreeze', 'beat breeze', 'bms', 'licence inclusive'],
    'SONOS': ['sonos'],
}

# Billing entity normalization (Company stores full names, finance uses shortcodes)
ENTITY_NORMALIZATION = {
    'BMAsia (Thailand) Co., Ltd.': 'bmasia_th',
    'BMAsia (Thailand) Co. Ltd.': 'bmasia_th',
    'BMAsia Thailand': 'bmasia_th',
    'bmasia_th': 'bmasia_th',
    'BMAsia Limited': 'bmasia_hk',
    'BMAsia Limited (HK)': 'bmasia_hk',
    'BMAsia HK': 'bmasia_hk',
    'bmasia_hk': 'bmasia_hk',
}

# Quarter date boundaries
QUARTER_DATES = {
    1: (1, 1, 3, 31),   # Jan 1 - Mar 31
    2: (4, 1, 6, 30),   # Apr 1 - Jun 30
    3: (7, 1, 9, 30),   # Jul 1 - Sep 30
    4: (10, 1, 12, 31), # Oct 1 - Dec 31
}


class RevenueRecognitionService:
    """Service for accrual-based revenue recognition calculations."""

    def __init__(self):
        from crm_app.models import (
            RevenueRecognitionSchedule,
            RevenueRecognitionEntry,
            Invoice,
            InvoiceLineItem,
        )
        self.Schedule = RevenueRecognitionSchedule
        self.Entry = RevenueRecognitionEntry
        self.Invoice = Invoice
        self.InvoiceLineItem = InvoiceLineItem

    # =========================================================================
    # CORE RECOGNITION CALCULATION
    # =========================================================================

    def calculate_quarterly_recognition(
        self,
        amount: Decimal,
        duration_months: Decimal,
        service_start: date,
        service_end: date,
        invoice_date: date,
        year: int,
        quarter: int
    ) -> Decimal:
        """
        Calculate recognized revenue for a specific quarter using Pom's exact formula.

        Verified against Pom's Excel:
            daily_rate = amount / total_service_days
            quarterly_income = daily_rate × service_days_in_quarter

        Service days = actual calendar days (inclusive).
        Revenue must NOT be recognized before invoice_date.
        """
        q_start_month, q_start_day, q_end_month, q_end_day = QUARTER_DATES[quarter]
        q_start = date(year, q_start_month, q_start_day)
        q_end = date(year, q_end_month, q_end_day)

        # Total service period days (inclusive)
        total_service_days = (service_end - service_start).days + 1
        if total_service_days <= 0:
            return Decimal('0')

        # Clamp to service period
        effective_start = max(q_start, service_start)
        effective_end = min(q_end, service_end)

        # No recognition before invoice date
        effective_start = max(effective_start, invoice_date)

        # No overlap = no recognition
        if effective_start > effective_end:
            return Decimal('0')

        # Quarter service days (inclusive)
        quarter_days = (effective_end - effective_start).days + 1

        if quarter_days <= 0:
            return Decimal('0')

        # Pom's formula: (amount / total_days) × quarter_days
        daily_rate = amount / Decimal(str(total_service_days))
        recognized = daily_rate * Decimal(str(quarter_days))

        return recognized.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    # =========================================================================
    # SCHEDULE + ENTRY GENERATION
    # =========================================================================

    def generate_entries_for_schedule(self, schedule, years: List[int] = None):
        """Generate all quarterly entries for a single schedule."""
        if years is None:
            years = list(range(
                schedule.service_period_start.year,
                schedule.service_period_end.year + 1
            ))

        entries = []
        cumulative = Decimal('0')

        # Pre-calculate cumulative from prior years (for correct balance when
        # generating a subset of years, e.g. years=[2025] for a 2024-2025 schedule)
        first_year = min(years)
        for prior_year in range(schedule.service_period_start.year, first_year):
            for q in range(1, 5):
                cumulative += self.calculate_quarterly_recognition(
                    amount=schedule.amount,
                    duration_months=schedule.duration_months,
                    service_start=schedule.service_period_start,
                    service_end=schedule.service_period_end,
                    invoice_date=schedule.invoice_date,
                    year=prior_year,
                    quarter=q,
                )

        for year in sorted(years):
            for quarter in range(1, 5):
                recognized = self.calculate_quarterly_recognition(
                    amount=schedule.amount,
                    duration_months=schedule.duration_months,
                    service_start=schedule.service_period_start,
                    service_end=schedule.service_period_end,
                    invoice_date=schedule.invoice_date,
                    year=year,
                    quarter=quarter
                )

                if recognized > 0 or cumulative > 0:
                    cumulative += recognized
                    balance = max(schedule.amount - cumulative, Decimal('0'))

                    entries.append(self.Entry(
                        schedule=schedule,
                        year=year,
                        quarter=quarter,
                        recognized_amount=recognized,
                        balance=balance.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
                    ))

        return entries

    @transaction.atomic
    def generate_entries_for_year(self, year: int, billing_entity: str):
        """Batch-generate all quarterly entries for a year + entity."""
        schedules = self.Schedule.objects.filter(
            billing_entity=billing_entity,
            status='active',
            service_period_start__year__lte=year,
            service_period_end__year__gte=year,
        )

        # Delete existing non-overridden entries for this year
        self.Entry.objects.filter(
            schedule__billing_entity=billing_entity,
            year=year,
            is_manually_overridden=False,
        ).delete()

        all_entries = []
        for schedule in schedules:
            entries = self.generate_entries_for_schedule(schedule, years=[year])
            # Skip entries that are manually overridden (already exist)
            for entry in entries:
                existing = self.Entry.objects.filter(
                    schedule=schedule,
                    year=entry.year,
                    quarter=entry.quarter,
                    is_manually_overridden=True,
                ).first()
                if not existing:
                    all_entries.append(entry)

        self.Entry.objects.bulk_create(all_entries, ignore_conflicts=True)
        logger.info(f"Generated {len(all_entries)} entries for {year} {billing_entity}")
        return len(all_entries)

    @transaction.atomic
    def regenerate_all_entries(self, year: int, billing_entity: str):
        """Full recalculation — deletes ALL entries (including overrides) and regenerates."""
        schedules = self.Schedule.objects.filter(
            billing_entity=billing_entity,
            status='active',
        )

        # Get all years that these schedules span
        years_needed = set()
        for s in schedules:
            for y in range(s.service_period_start.year, s.service_period_end.year + 1):
                years_needed.add(y)

        # If year specified, only do that year
        if year:
            years_needed = {year}

        self.Entry.objects.filter(
            schedule__billing_entity=billing_entity,
            year__in=years_needed,
        ).delete()

        all_entries = []
        for schedule in schedules:
            entries = self.generate_entries_for_schedule(schedule, years=sorted(years_needed))
            all_entries.extend(entries)

        self.Entry.objects.bulk_create(all_entries, ignore_conflicts=True)
        logger.info(f"Regenerated {len(all_entries)} entries for {billing_entity} years={sorted(years_needed)}")
        return len(all_entries)

    # =========================================================================
    # INVOICE → SCHEDULE AUTO-GENERATION
    # =========================================================================

    @transaction.atomic
    def generate_schedule_from_invoice(self, invoice) -> List:
        """Auto-create recognition schedules from an invoice's line items."""
        created = []
        # Determine billing entity from contract→company or direct company
        entity_name = ''
        if invoice.contract and invoice.contract.company:
            entity_name = invoice.contract.company.billing_entity or ''
        elif invoice.company:
            entity_name = invoice.company.billing_entity or ''
        entity = self.normalize_billing_entity(entity_name)

        for item in invoice.line_items.all():
            svc_start = item.service_period_start or invoice.service_period_start
            svc_end = item.service_period_end or invoice.service_period_end

            if not svc_start or not svc_end:
                logger.warning(f"Skipping line item {item.id} — no service period")
                continue

            duration = self._calc_duration_months(svc_start, svc_end)
            product = self.classify_product(item.product_service or item.description or '')

            schedule = self.Schedule.objects.create(
                invoice=invoice,
                invoice_line_item=item,
                invoice_number=invoice.invoice_number,
                invoice_date=invoice.issue_date,
                client_name=invoice.contract.company.name if invoice.contract and invoice.contract.company else (invoice.company.name if invoice.company else ''),
                billing_entity=entity,
                currency=invoice.currency or 'THB',
                product=product,
                amount=item.line_total or Decimal('0'),
                quantity=item.quantity or 1,
                sales_price=item.unit_price,
                service_period_start=svc_start,
                service_period_end=svc_end,
                duration_months=duration,
                is_imported=False,
            )

            # Generate entries for all years the service spans
            entries = self.generate_entries_for_schedule(schedule)
            self.Entry.objects.bulk_create(entries, ignore_conflicts=True)

            created.append(schedule)

        return created

    # =========================================================================
    # SUMMARY + REPORTING
    # =========================================================================

    def get_quarterly_summary(
        self,
        year: int,
        billing_entity: str,
        product: str = None,
        currency: str = None,
    ) -> Dict:
        """Get aggregated quarterly summary matching Pom's format."""
        filters = Q(
            schedule__billing_entity=billing_entity,
            schedule__status='active',
            year=year,
        )
        if product:
            filters &= Q(schedule__product=product)
        if currency:
            filters &= Q(schedule__currency=currency)

        entries = self.Entry.objects.filter(filters)

        # Per-quarter aggregation
        quarterly = {}
        for q in range(1, 5):
            q_entries = entries.filter(quarter=q)
            agg = q_entries.aggregate(
                total_income=Sum('recognized_amount'),
                total_balance=Sum('balance'),
            )
            quarterly[f'Q{q}'] = {
                'income': float(agg['total_income'] or 0),
                'balance': float(agg['total_balance'] or 0),
            }

        # Per-product breakdown
        products = {}
        for prod_code in ['SYB', 'LIM', 'SONOS', 'OTHER']:
            prod_entries = entries.filter(schedule__product=prod_code)
            if not prod_entries.exists():
                continue

            prod_quarterly = {}
            for q in range(1, 5):
                q_agg = prod_entries.filter(quarter=q).aggregate(
                    total_income=Sum('recognized_amount'),
                    total_balance=Sum('balance'),
                )
                prod_quarterly[f'Q{q}'] = {
                    'income': float(q_agg['total_income'] or 0),
                    'balance': float(q_agg['total_balance'] or 0),
                }
            products[prod_code] = prod_quarterly

        # Total amount from schedules
        schedule_filters = Q(
            billing_entity=billing_entity,
            status='active',
        )
        if product:
            schedule_filters &= Q(product=product)
        if currency:
            schedule_filters &= Q(currency=currency)

        # Only schedules that overlap with this year
        schedule_filters &= Q(
            service_period_start__year__lte=year,
            service_period_end__year__gte=year,
        )

        total_amount = self.Schedule.objects.filter(schedule_filters).aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0')

        total_recognized = entries.aggregate(
            total=Sum('recognized_amount')
        )['total'] or Decimal('0')

        schedule_count = self.Schedule.objects.filter(schedule_filters).count()

        return {
            'year': year,
            'billing_entity': billing_entity,
            'total_invoice_amount': float(total_amount),
            'total_recognized_ytd': float(total_recognized),
            'total_deferred': float(total_amount - total_recognized),
            'recognition_pct': float(
                (total_recognized / total_amount * 100) if total_amount > 0 else 0
            ),
            'schedule_count': schedule_count,
            'quarterly': quarterly,
            'by_product': products,
        }

    def get_deferred_revenue_balance(
        self,
        year: int,
        quarter: int,
        billing_entity: str,
        currency: str = None,
    ) -> Decimal:
        """Get single deferred revenue number for balance sheet integration."""
        filters = Q(
            schedule__billing_entity=billing_entity,
            schedule__status='active',
            year=year,
            quarter=quarter,
        )
        if currency:
            filters &= Q(schedule__currency=currency)

        result = self.Entry.objects.filter(filters).aggregate(
            total_balance=Sum('balance')
        )
        return Decimal(str(result['total_balance'] or 0))

    def get_schedules_detail(
        self,
        year: int,
        billing_entity: str,
        product: str = None,
        currency: str = None,
        status: str = 'active',
    ) -> List[Dict]:
        """Get detailed schedule list with quarterly entries — matches Pom's row format."""
        filters = Q(
            billing_entity=billing_entity,
            service_period_start__year__lte=year,
            service_period_end__year__gte=year,
        )
        if status:
            filters &= Q(status=status)
        if product:
            filters &= Q(product=product)
        if currency:
            filters &= Q(currency=currency)

        schedules = self.Schedule.objects.filter(filters).select_related(
            'invoice', 'invoice_line_item'
        ).prefetch_related('entries')

        rows = []
        for s in schedules:
            row = {
                'id': s.id,
                'invoice_number': s.invoice_number,
                'invoice_date': s.invoice_date.isoformat(),
                'client_name': s.client_name,
                'memo': s.memo,
                'product': s.product,
                'revenue_class': s.revenue_class,
                'currency': s.currency,
                'quantity': float(s.quantity),
                'sales_price': float(s.sales_price) if s.sales_price else None,
                'amount': float(s.amount),
                'service_period_start': s.service_period_start.isoformat(),
                'service_period_end': s.service_period_end.isoformat(),
                'duration_months': float(s.duration_months),
                'status': s.status,
                'is_imported': s.is_imported,
            }

            # Build lookup from prefetched entries (no extra DB queries)
            entry_map = {}
            for e in s.entries.all():
                if e.year == year:
                    entry_map[e.quarter] = e

            for q in range(1, 5):
                entry = entry_map.get(q)
                row[f'q{q}_income'] = float(entry.recognized_amount) if entry else 0
                row[f'q{q}_balance'] = float(entry.balance) if entry else float(s.amount)

            rows.append(row)

        return rows

    # =========================================================================
    # EXCEL IMPORT (Pom's format)
    # =========================================================================

    @transaction.atomic
    def import_from_excel(self, file_obj, billing_entity: str, currency: str) -> Dict:
        """
        Parse Pom's Excel → create schedules + entries.

        Verified column layout (from actual Excel files):
        C1=Type, C2=Date, C3=Num, C4=Name, C5=Memo, C6=Item, C7=Class,
        C8=Qty, C9=Sales Price, C10=Amount, C11=Service Period (text),
        C12=Start Date (dd/mm/yyyy text), C13=End Date (text),
        C14=Start Date (datetime), C15=End Date (datetime),
        C18=Total Days, C19=Q1 Days, C20=INCOME Q1, C21=BALANCE Q1,
        C22=Q2 Days, C23=INCOME Q2, C24=BALANCE Q2,
        C25=Q3 Days, C26=INCOME Q3, C27=BALANCE Q3,
        C28=Q4 Days, C29=INCOME Q4, C30=BALANCE Q4

        Returns dict with import stats.
        """
        import openpyxl
        from io import BytesIO

        wb = openpyxl.load_workbook(BytesIO(file_obj.read()), data_only=True)
        ws = wb.active

        # Find header row — look for 'Type' in column 1
        header_row = None
        for row_idx in range(1, 11):
            val = ws.cell(row=row_idx, column=1).value
            if val and str(val).strip().lower() == 'type':
                header_row = row_idx
                break

        if not header_row:
            # Fallback: scan all columns
            for row_idx in range(1, 11):
                for col_idx in range(1, 20):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val and str(val).strip().lower() == 'type':
                        header_row = row_idx
                        break
                if header_row:
                    break

        if not header_row:
            return {'error': 'Could not find header row (looking for "Type" in first 10 rows)', 'created': 0, 'skipped': 0}

        # Build header map from the actual header row
        headers = {}
        for col_idx in range(1, 50):
            val = ws.cell(row=header_row, column=col_idx).value
            if val:
                headers[str(val).strip()] = col_idx

        logger.info(f"Import: header_row={header_row}, headers={list(headers.keys())}")

        # Build column mapping — try header-based first, fall back to positional
        col_map = self._build_column_map(headers, header_row)

        # Detect the year from the header (e.g., "INCOME Q1/2025" → 2025)
        import_year = self._detect_year_from_headers(headers)

        stats = {'created': 0, 'skipped': 0, 'errors': [], 'total_amount': Decimal('0')}
        schedules_to_create = []
        quarterly_data_list = []  # parallel list — avoids relying on object attrs after bulk_create

        for row_idx in range(header_row + 1, ws.max_row + 1):
            try:
                # Read Type column
                row_type = ws.cell(row=row_idx, column=col_map['type']).value
                if not row_type:
                    continue

                type_str = str(row_type).strip().lower()

                # Skip non-data rows (totals, subtotals, blank)
                if type_str in ('total', 'subtotal', ''):
                    continue

                # Read Amount
                amount = self._parse_decimal(ws.cell(row=row_idx, column=col_map['amount']).value)
                if amount is None:
                    stats['skipped'] += 1
                    continue

                # Credit notes have negative amounts — include them
                if amount == 0:
                    stats['skipped'] += 1
                    continue

                # Parse dates — try datetime columns first (C14/C15), then text (C12/C13)
                start_date = None
                end_date = None

                if col_map.get('start_date_dt'):
                    start_date = self._parse_date(ws.cell(row=row_idx, column=col_map['start_date_dt']).value)
                if col_map.get('end_date_dt'):
                    end_date = self._parse_date(ws.cell(row=row_idx, column=col_map['end_date_dt']).value)

                if not start_date and col_map.get('start_date'):
                    start_date = self._parse_date(ws.cell(row=row_idx, column=col_map['start_date']).value)
                if not end_date and col_map.get('end_date'):
                    end_date = self._parse_date(ws.cell(row=row_idx, column=col_map['end_date']).value)

                inv_date = self._parse_date(ws.cell(row=row_idx, column=col_map['date']).value)

                if not inv_date or not start_date or not end_date:
                    stats['skipped'] += 1
                    continue

                # Duration in months (for display)
                duration = self._calc_duration_months(start_date, end_date)

                # Product classification
                item_val = ws.cell(row=row_idx, column=col_map['item']).value if col_map.get('item') else ''
                product = self.classify_product(str(item_val or ''))

                # Revenue class
                class_val = ws.cell(row=row_idx, column=col_map['class']).value if col_map.get('class') else ''
                revenue_class = self._normalize_revenue_class(str(class_val or ''))

                # Other fields
                inv_num = str(ws.cell(row=row_idx, column=col_map['num']).value or '').strip()
                client_name = str(ws.cell(row=row_idx, column=col_map['name']).value or '').strip()
                memo = str(ws.cell(row=row_idx, column=col_map.get('memo', 5)).value or '').strip()[:500]
                qty = self._parse_decimal(ws.cell(row=row_idx, column=col_map['qty']).value) or Decimal('1')
                sales_price = self._parse_decimal(ws.cell(row=row_idx, column=col_map['sales_price']).value)

                schedule = self.Schedule(
                    invoice_number=inv_num,
                    invoice_date=inv_date,
                    client_name=client_name,
                    memo=memo,
                    billing_entity=billing_entity,
                    currency=currency,
                    product=product,
                    revenue_class=revenue_class,
                    amount=amount,
                    quantity=qty,
                    sales_price=sales_price,
                    service_period_start=start_date,
                    service_period_end=end_date,
                    duration_months=duration,
                    status='active',
                    is_imported=True,
                )
                schedules_to_create.append(schedule)
                stats['total_amount'] += amount

                # Parse quarterly income/balance from Excel
                quarterly_data = {}
                for q in range(1, 5):
                    inc_col = col_map.get(f'income_q{q}')
                    bal_col = col_map.get(f'balance_q{q}')
                    if inc_col:
                        inc_val = self._parse_decimal(ws.cell(row=row_idx, column=inc_col).value)
                        bal_val = self._parse_decimal(ws.cell(row=row_idx, column=bal_col).value) if bal_col else None
                        if inc_val is not None or bal_val is not None:
                            quarterly_data[q] = {
                                'income': inc_val or Decimal('0'),
                                'balance': bal_val or Decimal('0'),
                            }

                # Store in parallel list (not on object — bulk_create may lose attrs)
                quarterly_data_list.append((quarterly_data, import_year))

            except Exception as e:
                stats['errors'].append(f"Row {row_idx}: {str(e)}")
                stats['skipped'] += 1

        # Bulk create schedules
        created_schedules = self.Schedule.objects.bulk_create(schedules_to_create)
        stats['created'] = len(created_schedules)

        # Create entries — use imported quarterly data if available, else calculate
        entries_to_create = []
        for i, schedule in enumerate(created_schedules):
            quarterly_data, yr = quarterly_data_list[i]
            yr = yr or schedule.service_period_start.year

            if quarterly_data:
                # Use Pom's exact numbers from the Excel
                for q, data in quarterly_data.items():
                    entries_to_create.append(self.Entry(
                        schedule=schedule,
                        year=yr,
                        quarter=q,
                        recognized_amount=data['income'],
                        balance=data['balance'],
                    ))
            else:
                # Calculate from scratch
                entries = self.generate_entries_for_schedule(schedule)
                entries_to_create.extend(entries)

        self.Entry.objects.bulk_create(entries_to_create, ignore_conflicts=True)
        stats['entries_created'] = len(entries_to_create)
        stats['total_amount'] = float(stats['total_amount'])

        logger.info(
            f"Import complete: {stats['created']} schedules, "
            f"{stats['entries_created']} entries, "
            f"total {currency} {stats['total_amount']}"
        )
        return stats

    # =========================================================================
    # MODIFICATION HANDLERS
    # =========================================================================

    @transaction.atomic
    def handle_cancellation(self, schedule_id: int) -> Dict:
        """Cancel a schedule and remove future entries."""
        schedule = self.Schedule.objects.get(id=schedule_id)
        schedule.status = 'cancelled'
        schedule.save()

        # Remove entries from future quarters
        today = date.today()
        current_quarter = (today.month - 1) // 3 + 1
        deleted_count = self.Entry.objects.filter(
            schedule=schedule,
        ).filter(
            Q(year__gt=today.year) |
            Q(year=today.year, quarter__gt=current_quarter)
        ).delete()[0]

        return {
            'schedule_id': schedule_id,
            'status': 'cancelled',
            'entries_removed': deleted_count,
        }

    @transaction.atomic
    def handle_modification(
        self,
        schedule_id: int,
        new_amount: Decimal = None,
        new_end_date: date = None,
    ) -> Dict:
        """Modify a schedule and recalculate remaining entries."""
        schedule = self.Schedule.objects.get(id=schedule_id)
        schedule.status = 'modified'

        if new_amount is not None:
            schedule.amount = new_amount
        if new_end_date is not None:
            schedule.service_period_end = new_end_date
            schedule.duration_months = self._calc_duration_months(
                schedule.service_period_start, new_end_date
            )

        schedule.save()

        # Regenerate all entries for this schedule
        self.Entry.objects.filter(
            schedule=schedule,
            is_manually_overridden=False,
        ).delete()

        entries = self.generate_entries_for_schedule(schedule)
        self.Entry.objects.bulk_create(entries, ignore_conflicts=True)

        return {
            'schedule_id': schedule_id,
            'status': 'modified',
            'entries_regenerated': len(entries),
        }

    # =========================================================================
    # HELPER METHODS
    # =========================================================================

    @staticmethod
    def classify_product(text: str) -> str:
        """Classify product from text description."""
        if not text:
            return 'OTHER'
        text_lower = text.lower().strip()

        for product, keywords in PRODUCT_KEYWORDS.items():
            for kw in keywords:
                if kw in text_lower:
                    return product

        # Direct match on product codes
        text_upper = text.upper().strip()
        if text_upper in ('SYB', 'LIM', 'SONOS'):
            return text_upper

        return 'OTHER'

    @staticmethod
    def normalize_billing_entity(entity_name: str) -> str:
        """Normalize company entity name to shortcode."""
        if not entity_name:
            return 'bmasia_th'

        # Direct lookup
        normalized = ENTITY_NORMALIZATION.get(entity_name.strip())
        if normalized:
            return normalized

        # Fuzzy match
        lower = entity_name.lower()
        if 'thailand' in lower or 'thai' in lower or 'th' in lower:
            return 'bmasia_th'
        if 'limited' in lower or 'hk' in lower or 'hong kong' in lower:
            return 'bmasia_hk'

        return 'bmasia_th'

    @staticmethod
    def _calc_duration_months(start: date, end: date) -> Decimal:
        """Calculate duration in months between two dates."""
        months = (end.year - start.year) * 12 + (end.month - start.month)
        # Add partial month if end.day > start.day
        if end.day >= start.day:
            months += Decimal(str(end.day - start.day + 1)) / Decimal('30')
        else:
            months += Decimal('1')  # Round up partial month
        return max(Decimal(str(months)).quantize(Decimal('0.01')), Decimal('0.01'))

    @staticmethod
    def _parse_date(value) -> Optional[date]:
        """Parse date from various formats."""
        if isinstance(value, date):
            return value
        if hasattr(value, 'date'):
            return value.date()
        if not value:
            return None
        try:
            from datetime import datetime
            s = str(value).strip()
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%d %b %Y', '%B %d, %Y'):
                try:
                    return datetime.strptime(s, fmt).date()
                except ValueError:
                    continue
        except Exception:
            pass
        return None

    @staticmethod
    def _parse_decimal(value) -> Optional[Decimal]:
        """Parse decimal from various formats."""
        if value is None:
            return None
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        try:
            # Remove commas and currency symbols
            s = str(value).strip().replace(',', '').replace('$', '').replace('฿', '')
            if not s or s == '-' or s == '':
                return None
            return Decimal(s)
        except Exception:
            return None

    def _build_column_map(self, headers: Dict[str, int], header_row: int) -> Dict[str, int]:
        """
        Build column mapping from Pom's Excel headers.
        Uses header text matching first, then falls back to known positional layout.
        """
        mapping = {}

        # Standard column aliases
        aliases = {
            'type': ['Type', 'type'],
            'date': ['Date', 'date'],
            'num': ['Num', 'num', 'Number'],
            'name': ['Name', 'name'],
            'memo': ['Memo', 'memo'],
            'item': ['Item', 'item'],
            'class': ['Class', 'class'],
            'qty': ['Qty', 'qty', 'QTY'],
            'sales_price': ['Sales Price', 'sales_price'],
            'amount': ['Amount', 'amount', 'Net Amount', 'Net amount'],
        }

        for key, header_names in aliases.items():
            for h in header_names:
                if h in headers:
                    mapping[key] = headers[h]
                    break

        # Fall back to positional (Pom's standard layout: C1-C10)
        positional_defaults = {
            'type': 1, 'date': 2, 'num': 3, 'name': 4, 'memo': 5,
            'item': 6, 'class': 7, 'qty': 8, 'sales_price': 9, 'amount': 10,
        }
        for key, col in positional_defaults.items():
            if key not in mapping:
                mapping[key] = col

        # Start/End date columns — varies between files:
        # HK file: C12=text start, C13=text end, C14=datetime start, C15=datetime end
        # TH file: C12=text start, C13=text end, C14=mm/yyyy, C15=mm/yyyy (no datetime)
        # We scan for columns with 'dd/mm/yyyy' in header → those are start/end text dates
        date_cols = []
        for h_name, col_idx in sorted(headers.items(), key=lambda x: x[1]):
            h_str = str(h_name).lower().strip() if h_name else ''
            if 'dd/mm/yyyy' in h_str or 'stdd' in h_str or 'eddd' in h_str:
                date_cols.append(col_idx)

        if len(date_cols) >= 2:
            mapping['start_date'] = date_cols[0]
            mapping['end_date'] = date_cols[1]
        else:
            mapping.setdefault('start_date', 12)
            mapping.setdefault('end_date', 13)

        # HK file has datetime versions at C14/C15 — check if they contain actual dates
        mapping['start_date_dt'] = 14
        mapping['end_date_dt'] = 15

        # Quarterly INCOME/BALANCE columns — match by header text
        # Note: TH file has 2 years (2024 + 2025 columns). We take the FIRST
        # set of Q1-Q4 columns found (leftmost), which matches the primary data year.
        for h_name, col_idx in sorted(headers.items(), key=lambda x: x[1]):
            if not h_name:
                continue
            h_lower = h_name.lower().strip()
            for q in range(1, 5):
                inc_key = f'income_q{q}'
                bal_key = f'balance_q{q}'
                if ('income' in h_lower) and (f'q{q}' in h_lower) and inc_key not in mapping:
                    mapping[inc_key] = col_idx
                elif ('balance' in h_lower) and (f'q{q}' in h_lower) and bal_key not in mapping:
                    mapping[bal_key] = col_idx

        logger.info(f"Column mapping: {mapping}")
        return mapping

    def _detect_year_from_headers(self, headers: Dict[str, int]) -> Optional[int]:
        """Detect year from header text like 'INCOME Q1/2025'."""
        import re
        for h_name in headers.keys():
            if not h_name:
                continue
            match = re.search(r'20\d{2}', str(h_name))
            if match:
                return int(match.group())
        return None

    @staticmethod
    def _normalize_revenue_class(value: str) -> str:
        """Normalize revenue class to valid choices."""
        if not value:
            return 'Other'
        valid_classes = {'New', 'Renew', 'Add Outlet', 'Upgrade', 'Other'}
        # Direct match
        value = value.strip()
        if value in valid_classes:
            return value
        # Fuzzy match
        lower = value.lower()
        if 'new' in lower and 'renew' not in lower:
            return 'New'
        if 'renew' in lower:
            return 'Renew'
        if 'add' in lower or 'outlet' in lower:
            return 'Add Outlet'
        if 'upgrade' in lower:
            return 'Upgrade'
        return 'Other'
