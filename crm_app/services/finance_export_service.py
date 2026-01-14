"""
Finance Export Service for BMAsia CRM

Generates PDF and Excel exports for Finance Module reports:
- Profit & Loss Statement
- Cash Flow Statement
- Balance Sheet

Uses ReportLab for PDF and openpyxl for Excel generation.
Part of the Finance & Accounting Module - Phase 7.

Usage:
    from crm_app.services.finance_export_service import FinanceExportService

    service = FinanceExportService()
    pdf_buffer = service.generate_pl_pdf(data, year=2026, month=1, billing_entity='bmasia_th', currency='THB')
    excel_buffer = service.generate_pl_excel(data, year=2026, month=1, billing_entity='bmasia_th', currency='THB')
"""

import logging
from io import BytesIO
from datetime import datetime
from decimal import Decimal
from typing import Dict, Optional

# ReportLab imports for PDF
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image,
    HRFlowable, PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

# openpyxl imports for Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

import os
from django.conf import settings

logger = logging.getLogger(__name__)


# Billing entity information
BILLING_ENTITY_INFO = {
    'bmasia_th': {
        'name': 'BMAsia (Thailand) Co., Ltd.',
        'address': '725 S-Metro Building, Suite 144, Level 20, Sukhumvit Road, Klongtan Nuea Watthana, Bangkok 10110, Thailand',
        'phone': '+66 2153 3520',
        'tax_id': '0105548025073',
    },
    'bmasia_hk': {
        'name': 'BMAsia Limited',
        'address': '22nd Floor, Tai Yau Building, 181 Johnston Road, Wanchai, Hong Kong',
        'phone': '+66 2153 3520',
        'tax_id': None,
    },
    'all': {
        'name': 'BMAsia (Consolidated)',
        'address': '',
        'phone': '+66 2153 3520',
        'tax_id': None,
    }
}

# Currency symbols
CURRENCY_SYMBOLS = {
    'THB': '฿',
    'USD': '$',
    'EUR': '€',
    'GBP': '£',
    'HKD': 'HK$',
    'all': '',
}


class FinanceExportService:
    """
    Service for generating PDF and Excel exports of financial reports.
    """

    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()

    def _setup_custom_styles(self):
        """Set up custom PDF styles for BMAsia branding."""
        # Title style
        self.title_style = ParagraphStyle(
            'FinanceTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#FFA500'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )

        # Report heading style
        self.heading_style = ParagraphStyle(
            'FinanceHeading',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#424242'),
            spaceAfter=8,
            spaceBefore=12,
            fontName='Helvetica-Bold'
        )

        # Section heading style
        self.section_style = ParagraphStyle(
            'SectionHeading',
            parent=self.styles['Heading3'],
            fontSize=11,
            textColor=colors.HexColor('#424242'),
            spaceAfter=4,
            spaceBefore=8,
            fontName='Helvetica-Bold'
        )

        # Body text style
        self.body_style = ParagraphStyle(
            'FinanceBody',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#424242'),
            leading=14
        )

        # Small/footer style
        self.small_style = ParagraphStyle(
            'FinanceSmall',
            parent=self.styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#757575'),
            leading=10
        )

        # Right-aligned number style
        self.number_style = ParagraphStyle(
            'FinanceNumber',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#424242'),
            alignment=TA_RIGHT
        )

    def _get_entity_info(self, billing_entity: str) -> Dict:
        """Get billing entity information."""
        return BILLING_ENTITY_INFO.get(billing_entity, BILLING_ENTITY_INFO['all'])

    def _get_currency_symbol(self, currency: str) -> str:
        """Get currency symbol."""
        return CURRENCY_SYMBOLS.get(currency, currency + ' ' if currency else '')

    def _format_currency(self, value: float, currency: str) -> str:
        """Format a value with currency symbol."""
        symbol = self._get_currency_symbol(currency)
        if value < 0:
            return f"({symbol}{abs(value):,.2f})"
        return f"{symbol}{value:,.2f}"

    def _format_percentage(self, value: float) -> str:
        """Format a percentage value."""
        return f"{value:.1f}%"

    def _add_pdf_header(self, elements: list, title: str, period_text: str,
                        billing_entity: str, currency: str):
        """Add standard header to PDF with logo and title."""
        # Logo
        logo_path = os.path.join(
            settings.BASE_DIR, 'crm_app', 'static', 'crm_app', 'images', 'bmasia_logo.png'
        )
        try:
            if os.path.exists(logo_path):
                logo = Image(logo_path, width=140, height=56, kind='proportional')
                logo.hAlign = 'LEFT'
                elements.append(logo)
            else:
                elements.append(Paragraph("BM ASIA", self.title_style))
        except Exception:
            elements.append(Paragraph("BM ASIA", self.title_style))

        # Orange accent line
        elements.append(Spacer(1, 0.05 * inch))
        elements.append(HRFlowable(
            width="100%", thickness=2,
            color=colors.HexColor('#FFA500'),
            spaceBefore=0, spaceAfter=0
        ))
        elements.append(Spacer(1, 0.15 * inch))

        # Report title
        elements.append(Paragraph(title, self.title_style))

        # Period and entity info
        entity_info = self._get_entity_info(billing_entity)
        currency_display = currency if currency != 'all' else 'All Currencies'

        info_text = f"""
        <b>Period:</b> {period_text}<br/>
        <b>Entity:</b> {entity_info['name']}<br/>
        <b>Currency:</b> {currency_display}
        """
        elements.append(Paragraph(info_text, self.body_style))
        elements.append(Spacer(1, 0.2 * inch))

    def _add_pdf_footer(self, elements: list):
        """Add footer with generation timestamp."""
        elements.append(Spacer(1, 0.3 * inch))
        elements.append(HRFlowable(
            width="100%", thickness=0.5,
            color=colors.HexColor('#e0e0e0'),
            spaceBefore=0, spaceAfter=0
        ))
        elements.append(Spacer(1, 0.1 * inch))

        footer_text = f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        elements.append(Paragraph(footer_text, self.small_style))

    def _create_table_style(self, has_subtotals: bool = False) -> TableStyle:
        """Create standard table style with BMAsia branding."""
        style_commands = [
            # Header row - orange background
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFA500')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),

            # Data rows
            ('FONT', (0, 1), (-1, -1), 'Helvetica', 9),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#424242')),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),

            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e0')),

            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fafafa')]),
        ]

        return TableStyle(style_commands)

    # =========================================================================
    # PROFIT & LOSS EXPORTS
    # =========================================================================

    def generate_pl_pdf(
        self,
        data: Dict,
        year: int,
        month: int = None,
        billing_entity: str = 'all',
        currency: str = 'all'
    ) -> BytesIO:
        """
        Generate PDF for Profit & Loss Statement.

        Args:
            data: P&L data from ProfitLossService
            year: Year
            month: Month (None for YTD)
            billing_entity: Billing entity filter
            currency: Currency filter

        Returns:
            BytesIO buffer containing the PDF
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=letter,
            topMargin=0.5 * inch, bottomMargin=0.5 * inch,
            leftMargin=0.5 * inch, rightMargin=0.5 * inch
        )

        elements = []

        # Determine period text
        if data.get('period', {}).get('type') == 'ytd':
            through_month = data['period'].get('through_month', month)
            month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                          'July', 'August', 'September', 'October', 'November', 'December']
            period_text = f"Year-to-Date through {month_names[through_month]} {year}"
        else:
            month_name = data.get('period', {}).get('month_name', '')
            period_text = f"{month_name} {year}"

        # Header
        self._add_pdf_header(
            elements, "PROFIT & LOSS STATEMENT",
            period_text, billing_entity, currency
        )

        curr = currency if currency != 'all' else 'THB'

        # Revenue Section
        elements.append(Paragraph("REVENUE", self.heading_style))

        revenue = data.get('revenue', {})
        revenue_data = [
            ['Category', 'Count', 'Amount'],
            ['New Contracts', str(revenue.get('new', {}).get('count', 0)),
             self._format_currency(revenue.get('new', {}).get('value', 0), curr)],
            ['Renewals', str(revenue.get('renewal', {}).get('count', 0)),
             self._format_currency(revenue.get('renewal', {}).get('value', 0), curr)],
            ['Add-ons', str(revenue.get('addon', {}).get('count', 0)),
             self._format_currency(revenue.get('addon', {}).get('value', 0), curr)],
            ['Churn', str(revenue.get('churn', {}).get('count', 0)),
             self._format_currency(revenue.get('churn', {}).get('value', 0), curr)],
            ['TOTAL REVENUE', '', self._format_currency(revenue.get('total', 0), curr)],
        ]

        revenue_table = Table(revenue_data, colWidths=[3.5 * inch, 1.5 * inch, 2 * inch])
        revenue_table.setStyle(self._create_table_style())
        # Bold the total row
        revenue_table.setStyle(TableStyle([
            ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 10),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fff3e0')),
        ]))
        elements.append(revenue_table)
        elements.append(Spacer(1, 0.2 * inch))

        # Cost of Goods Sold Section
        elements.append(Paragraph("COST OF GOODS SOLD", self.heading_style))

        cogs = data.get('cogs', {})
        cogs_data = [['Category', 'Amount']]
        for cat in cogs.get('categories', []):
            cogs_data.append([
                cat.get('category_name', 'Unknown'),
                self._format_currency(cat.get('amount', 0), curr)
            ])
        cogs_data.append(['TOTAL COGS', self._format_currency(cogs.get('total', 0), curr)])

        cogs_table = Table(cogs_data, colWidths=[5 * inch, 2 * inch])
        cogs_table.setStyle(self._create_table_style())
        cogs_table.setStyle(TableStyle([
            ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 10),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fff3e0')),
        ]))
        elements.append(cogs_table)
        elements.append(Spacer(1, 0.15 * inch))

        # Gross Profit
        gross_profit = data.get('gross_profit', 0)
        gross_margin = data.get('gross_margin', 0)
        elements.append(Paragraph(
            f"<b>GROSS PROFIT:</b> {self._format_currency(gross_profit, curr)} "
            f"({self._format_percentage(gross_margin)} margin)",
            self.body_style
        ))
        elements.append(Spacer(1, 0.2 * inch))

        # Operating Expenses Section
        elements.append(Paragraph("OPERATING EXPENSES", self.heading_style))

        opex = data.get('operating_expenses', {})

        # G&A Expenses
        elements.append(Paragraph("General & Administrative", self.section_style))
        gna = opex.get('gna', {})
        gna_data = [['Category', 'Amount']]
        for cat in gna.get('categories', []):
            gna_data.append([
                cat.get('category_name', 'Unknown'),
                self._format_currency(cat.get('amount', 0), curr)
            ])
        gna_data.append(['Subtotal G&A', self._format_currency(gna.get('total', 0), curr)])

        gna_table = Table(gna_data, colWidths=[5 * inch, 2 * inch])
        gna_table.setStyle(self._create_table_style())
        gna_table.setStyle(TableStyle([
            ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 9),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f5f5f5')),
        ]))
        elements.append(gna_table)
        elements.append(Spacer(1, 0.1 * inch))

        # Sales & Marketing Expenses
        elements.append(Paragraph("Sales & Marketing", self.section_style))
        sm = opex.get('sales_marketing', {})
        sm_data = [['Category', 'Amount']]
        for cat in sm.get('categories', []):
            sm_data.append([
                cat.get('category_name', 'Unknown'),
                self._format_currency(cat.get('amount', 0), curr)
            ])
        sm_data.append(['Subtotal S&M', self._format_currency(sm.get('total', 0), curr)])

        sm_table = Table(sm_data, colWidths=[5 * inch, 2 * inch])
        sm_table.setStyle(self._create_table_style())
        sm_table.setStyle(TableStyle([
            ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 9),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f5f5f5')),
        ]))
        elements.append(sm_table)
        elements.append(Spacer(1, 0.1 * inch))

        elements.append(Paragraph(
            f"<b>TOTAL OPERATING EXPENSES:</b> {self._format_currency(opex.get('total', 0), curr)}",
            self.body_style
        ))
        elements.append(Spacer(1, 0.2 * inch))

        # Summary Section
        elements.append(HRFlowable(
            width="100%", thickness=1,
            color=colors.HexColor('#FFA500'),
            spaceBefore=0, spaceAfter=0
        ))
        elements.append(Spacer(1, 0.15 * inch))

        operating_income = data.get('operating_income', 0)
        operating_margin = data.get('operating_margin', 0)
        net_profit = data.get('net_profit', 0)
        net_margin = data.get('net_margin', 0)

        summary_data = [
            ['OPERATING INCOME', self._format_currency(operating_income, curr),
             f"{self._format_percentage(operating_margin)} margin"],
            ['NET PROFIT', self._format_currency(net_profit, curr),
             f"{self._format_percentage(net_margin)} margin"],
        ]

        summary_table = Table(summary_data, colWidths=[3 * inch, 2.5 * inch, 1.5 * inch])
        summary_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica-Bold', 12),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#424242')),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f5e9') if net_profit >= 0 else colors.HexColor('#ffebee')),
        ]))
        elements.append(summary_table)

        # Footer
        self._add_pdf_footer(elements)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer

    def generate_pl_excel(
        self,
        data: Dict,
        year: int,
        month: int = None,
        billing_entity: str = 'all',
        currency: str = 'all'
    ) -> BytesIO:
        """
        Generate Excel for Profit & Loss Statement.

        Args:
            data: P&L data from ProfitLossService
            year: Year
            month: Month (None for YTD)
            billing_entity: Billing entity filter
            currency: Currency filter

        Returns:
            BytesIO buffer containing the Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Profit & Loss"

        # Styles
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        section_font = Font(bold=True, size=11)
        total_font = Font(bold=True, size=10)
        total_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        currency_format = '#,##0.00'
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        curr = currency if currency != 'all' else 'THB'
        entity_info = self._get_entity_info(billing_entity)

        # Determine period text
        if data.get('period', {}).get('type') == 'ytd':
            through_month = data['period'].get('through_month', month)
            month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                          'July', 'August', 'September', 'October', 'November', 'December']
            period_text = f"Year-to-Date through {month_names[through_month]} {year}"
        else:
            month_name = data.get('period', {}).get('month_name', '')
            period_text = f"{month_name} {year}"

        # Header section
        ws['A1'] = "PROFIT & LOSS STATEMENT"
        ws['A1'].font = Font(bold=True, size=16, color="FFA500")
        ws.merge_cells('A1:C1')

        ws['A2'] = f"Period: {period_text}"
        ws['A3'] = f"Entity: {entity_info['name']}"
        ws['A4'] = f"Currency: {curr}"
        ws['A5'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"

        row = 7

        # Revenue Section
        ws[f'A{row}'] = "REVENUE"
        ws[f'A{row}'].font = section_font
        row += 1

        ws[f'A{row}'] = "Category"
        ws[f'B{row}'] = "Count"
        ws[f'C{row}'] = "Amount"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].border = thin_border
        row += 1

        revenue = data.get('revenue', {})
        revenue_items = [
            ('New Contracts', revenue.get('new', {})),
            ('Renewals', revenue.get('renewal', {})),
            ('Add-ons', revenue.get('addon', {})),
            ('Churn', revenue.get('churn', {})),
        ]

        for name, item in revenue_items:
            ws[f'A{row}'] = name
            ws[f'B{row}'] = item.get('count', 0)
            ws[f'C{row}'] = item.get('value', 0)
            ws[f'C{row}'].number_format = currency_format
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = thin_border
            row += 1

        # Total Revenue
        ws[f'A{row}'] = "TOTAL REVENUE"
        ws[f'B{row}'] = ""
        ws[f'C{row}'] = revenue.get('total', 0)
        ws[f'C{row}'].number_format = currency_format
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = total_font
            ws[f'{col}{row}'].fill = total_fill
            ws[f'{col}{row}'].border = thin_border
        row += 2

        # COGS Section
        ws[f'A{row}'] = "COST OF GOODS SOLD"
        ws[f'A{row}'].font = section_font
        row += 1

        ws[f'A{row}'] = "Category"
        ws[f'B{row}'] = "Amount"
        for col in ['A', 'B']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].border = thin_border
        row += 1

        cogs = data.get('cogs', {})
        for cat in cogs.get('categories', []):
            ws[f'A{row}'] = cat.get('category_name', 'Unknown')
            ws[f'B{row}'] = cat.get('amount', 0)
            ws[f'B{row}'].number_format = currency_format
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = thin_border
            row += 1

        ws[f'A{row}'] = "TOTAL COGS"
        ws[f'B{row}'] = cogs.get('total', 0)
        ws[f'B{row}'].number_format = currency_format
        for col in ['A', 'B']:
            ws[f'{col}{row}'].font = total_font
            ws[f'{col}{row}'].fill = total_fill
            ws[f'{col}{row}'].border = thin_border
        row += 2

        # Gross Profit
        ws[f'A{row}'] = "GROSS PROFIT"
        ws[f'B{row}'] = data.get('gross_profit', 0)
        ws[f'C{row}'] = f"{data.get('gross_margin', 0):.1f}%"
        ws[f'A{row}'].font = total_font
        ws[f'B{row}'].font = total_font
        ws[f'B{row}'].number_format = currency_format
        row += 2

        # Operating Expenses
        ws[f'A{row}'] = "OPERATING EXPENSES"
        ws[f'A{row}'].font = section_font
        row += 1

        opex = data.get('operating_expenses', {})

        # G&A
        ws[f'A{row}'] = "General & Administrative"
        ws[f'A{row}'].font = Font(italic=True, bold=True)
        row += 1

        for cat in opex.get('gna', {}).get('categories', []):
            ws[f'A{row}'] = f"  {cat.get('category_name', 'Unknown')}"
            ws[f'B{row}'] = cat.get('amount', 0)
            ws[f'B{row}'].number_format = currency_format
            row += 1

        ws[f'A{row}'] = "  Subtotal G&A"
        ws[f'B{row}'] = opex.get('gna', {}).get('total', 0)
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].font = Font(italic=True)
        row += 1

        # S&M
        ws[f'A{row}'] = "Sales & Marketing"
        ws[f'A{row}'].font = Font(italic=True, bold=True)
        row += 1

        for cat in opex.get('sales_marketing', {}).get('categories', []):
            ws[f'A{row}'] = f"  {cat.get('category_name', 'Unknown')}"
            ws[f'B{row}'] = cat.get('amount', 0)
            ws[f'B{row}'].number_format = currency_format
            row += 1

        ws[f'A{row}'] = "  Subtotal S&M"
        ws[f'B{row}'] = opex.get('sales_marketing', {}).get('total', 0)
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].font = Font(italic=True)
        row += 1

        ws[f'A{row}'] = "TOTAL OPERATING EXPENSES"
        ws[f'B{row}'] = opex.get('total', 0)
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].font = total_font
        ws[f'B{row}'].font = total_font
        ws[f'A{row}'].fill = total_fill
        ws[f'B{row}'].fill = total_fill
        row += 2

        # Summary
        ws[f'A{row}'] = "OPERATING INCOME"
        ws[f'B{row}'] = data.get('operating_income', 0)
        ws[f'C{row}'] = f"{data.get('operating_margin', 0):.1f}%"
        ws[f'A{row}'].font = total_font
        ws[f'B{row}'].font = total_font
        ws[f'B{row}'].number_format = currency_format
        row += 1

        ws[f'A{row}'] = "NET PROFIT"
        ws[f'B{row}'] = data.get('net_profit', 0)
        ws[f'C{row}'] = f"{data.get('net_margin', 0):.1f}%"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'].number_format = currency_format

        net_profit = data.get('net_profit', 0)
        profit_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid") if net_profit >= 0 else PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        ws[f'A{row}'].fill = profit_fill
        ws[f'B{row}'].fill = profit_fill
        ws[f'C{row}'].fill = profit_fill

        # Auto-fit columns
        for col_letter in ['A', 'B', 'C']:
            ws.column_dimensions[col_letter].width = 25 if col_letter == 'A' else 18

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # =========================================================================
    # CASH FLOW EXPORTS
    # =========================================================================

    def generate_cash_flow_pdf(
        self,
        data: Dict,
        year: int,
        month: int = None,
        billing_entity: str = 'all',
        currency: str = 'all'
    ) -> BytesIO:
        """
        Generate PDF for Cash Flow Statement.

        Args:
            data: Cash Flow data from CashFlowService
            year: Year
            month: Month (None for YTD)
            billing_entity: Billing entity filter
            currency: Currency filter

        Returns:
            BytesIO buffer containing the PDF
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=letter,
            topMargin=0.5 * inch, bottomMargin=0.5 * inch,
            leftMargin=0.5 * inch, rightMargin=0.5 * inch
        )

        elements = []

        # Determine period text
        if data.get('period', {}).get('type') == 'ytd':
            through_month = data['period'].get('through_month', month)
            through_name = data['period'].get('through_month_name', '')
            period_text = f"Year-to-Date through {through_name} {year}"
        else:
            month_name = data.get('period', {}).get('month_name', '')
            period_text = f"{month_name} {year}"

        # Header
        self._add_pdf_header(
            elements, "CASH FLOW STATEMENT",
            period_text, billing_entity, currency
        )

        curr = currency if currency != 'all' else 'THB'

        # Operating Activities Section
        elements.append(Paragraph("OPERATING ACTIVITIES", self.heading_style))

        operating = data.get('operating_activities', {})
        operating_data = [
            ['Item', 'Amount'],
            ['Cash from Customers',
             self._format_currency(operating.get('cash_from_customers', {}).get('value', 0), curr)],
            ['Cash to Suppliers',
             f"({self._format_currency(operating.get('cash_to_suppliers', {}).get('value', 0), curr)})"],
            ['Cash to Employees',
             f"({self._format_currency(operating.get('cash_to_employees', {}).get('value', 0), curr)})"],
            ['Other Operating Cash',
             self._format_currency(operating.get('other', 0), curr)],
            ['NET CASH FROM OPERATIONS',
             self._format_currency(operating.get('net_cash_from_operations', 0), curr)],
        ]

        operating_table = Table(operating_data, colWidths=[5 * inch, 2 * inch])
        operating_table.setStyle(self._create_table_style())
        operating_table.setStyle(TableStyle([
            ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 10),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e3f2fd')),
        ]))
        elements.append(operating_table)
        elements.append(Spacer(1, 0.2 * inch))

        # Investing Activities Section
        elements.append(Paragraph("INVESTING ACTIVITIES", self.heading_style))

        investing = data.get('investing_activities', {})
        investing_data = [
            ['Item', 'Amount'],
            ['CapEx Purchases',
             f"({self._format_currency(investing.get('capex_purchases', {}).get('value', 0), curr)})"],
            ['Asset Sales',
             self._format_currency(investing.get('asset_sales', 0), curr)],
            ['Other Investing Cash',
             self._format_currency(investing.get('other', 0), curr)],
            ['NET CASH FROM INVESTING',
             self._format_currency(investing.get('net_cash_from_investing', 0), curr)],
        ]

        investing_table = Table(investing_data, colWidths=[5 * inch, 2 * inch])
        investing_table.setStyle(self._create_table_style())
        investing_table.setStyle(TableStyle([
            ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 10),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fce4ec')),
        ]))
        elements.append(investing_table)
        elements.append(Spacer(1, 0.2 * inch))

        # Financing Activities Section
        elements.append(Paragraph("FINANCING ACTIVITIES", self.heading_style))

        financing = data.get('financing_activities', {})
        financing_data = [
            ['Item', 'Amount'],
            ['Loan Proceeds',
             self._format_currency(financing.get('loan_proceeds', 0), curr)],
            ['Loan Repayments',
             f"({self._format_currency(financing.get('loan_repayments', 0), curr)})"],
            ['Equity Injections',
             self._format_currency(financing.get('equity_injections', 0), curr)],
            ['Dividends Paid',
             f"({self._format_currency(financing.get('dividends_paid', 0), curr)})"],
            ['Other Financing Cash',
             self._format_currency(financing.get('other', 0), curr)],
            ['NET CASH FROM FINANCING',
             self._format_currency(financing.get('net_cash_from_financing', 0), curr)],
        ]

        financing_table = Table(financing_data, colWidths=[5 * inch, 2 * inch])
        financing_table.setStyle(self._create_table_style())
        financing_table.setStyle(TableStyle([
            ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 10),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f3e5f5')),
        ]))
        elements.append(financing_table)
        elements.append(Spacer(1, 0.2 * inch))

        # Summary Section
        elements.append(HRFlowable(
            width="100%", thickness=1,
            color=colors.HexColor('#FFA500'),
            spaceBefore=0, spaceAfter=0
        ))
        elements.append(Spacer(1, 0.15 * inch))

        net_change = data.get('net_change_in_cash', 0)
        opening = data.get('opening_cash_balance', 0)
        closing = data.get('closing_cash_balance', 0)

        summary_data = [
            ['Description', 'Amount'],
            ['Net Change in Cash', self._format_currency(net_change, curr)],
            ['Opening Cash Balance', self._format_currency(opening, curr)],
            ['CLOSING CASH BALANCE', self._format_currency(closing, curr)],
        ]

        summary_table = Table(summary_data, colWidths=[5 * inch, 2 * inch])
        summary_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFA500')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONT', (0, 1), (-1, -2), 'Helvetica', 10),
            ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f5e9') if closing >= 0 else colors.HexColor('#ffebee')),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e0')),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(summary_table)

        # Footer
        self._add_pdf_footer(elements)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer

    def generate_cash_flow_excel(
        self,
        data: Dict,
        year: int,
        month: int = None,
        billing_entity: str = 'all',
        currency: str = 'all'
    ) -> BytesIO:
        """
        Generate Excel for Cash Flow Statement.

        Args:
            data: Cash Flow data from CashFlowService
            year: Year
            month: Month (None for YTD)
            billing_entity: Billing entity filter
            currency: Currency filter

        Returns:
            BytesIO buffer containing the Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Cash Flow"

        # Styles
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        section_font = Font(bold=True, size=11)
        total_font = Font(bold=True, size=10)
        currency_format = '#,##0.00'
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        curr = currency if currency != 'all' else 'THB'
        entity_info = self._get_entity_info(billing_entity)

        # Determine period text
        if data.get('period', {}).get('type') == 'ytd':
            through_name = data['period'].get('through_month_name', '')
            period_text = f"Year-to-Date through {through_name} {year}"
        else:
            month_name = data.get('period', {}).get('month_name', '')
            period_text = f"{month_name} {year}"

        # Header section
        ws['A1'] = "CASH FLOW STATEMENT"
        ws['A1'].font = Font(bold=True, size=16, color="FFA500")
        ws.merge_cells('A1:B1')

        ws['A2'] = f"Period: {period_text}"
        ws['A3'] = f"Entity: {entity_info['name']}"
        ws['A4'] = f"Currency: {curr}"
        ws['A5'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"

        row = 7

        # Operating Activities
        ws[f'A{row}'] = "OPERATING ACTIVITIES"
        ws[f'A{row}'].font = section_font
        row += 1

        ws[f'A{row}'] = "Item"
        ws[f'B{row}'] = "Amount"
        for col in ['A', 'B']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].border = thin_border
        row += 1

        operating = data.get('operating_activities', {})
        op_items = [
            ('Cash from Customers', operating.get('cash_from_customers', {}).get('value', 0)),
            ('Cash to Suppliers', -operating.get('cash_to_suppliers', {}).get('value', 0)),
            ('Cash to Employees', -operating.get('cash_to_employees', {}).get('value', 0)),
            ('Other Operating Cash', operating.get('other', 0)),
        ]

        for name, value in op_items:
            ws[f'A{row}'] = name
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = currency_format
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = thin_border
            row += 1

        ws[f'A{row}'] = "NET CASH FROM OPERATIONS"
        ws[f'B{row}'] = operating.get('net_cash_from_operations', 0)
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].font = total_font
        ws[f'B{row}'].font = total_font
        op_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        ws[f'A{row}'].fill = op_fill
        ws[f'B{row}'].fill = op_fill
        for col in ['A', 'B']:
            ws[f'{col}{row}'].border = thin_border
        row += 2

        # Investing Activities
        ws[f'A{row}'] = "INVESTING ACTIVITIES"
        ws[f'A{row}'].font = section_font
        row += 1

        ws[f'A{row}'] = "Item"
        ws[f'B{row}'] = "Amount"
        for col in ['A', 'B']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].border = thin_border
        row += 1

        investing = data.get('investing_activities', {})
        inv_items = [
            ('CapEx Purchases', -investing.get('capex_purchases', {}).get('value', 0)),
            ('Asset Sales', investing.get('asset_sales', 0)),
            ('Other Investing Cash', investing.get('other', 0)),
        ]

        for name, value in inv_items:
            ws[f'A{row}'] = name
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = currency_format
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = thin_border
            row += 1

        ws[f'A{row}'] = "NET CASH FROM INVESTING"
        ws[f'B{row}'] = investing.get('net_cash_from_investing', 0)
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].font = total_font
        ws[f'B{row}'].font = total_font
        inv_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
        ws[f'A{row}'].fill = inv_fill
        ws[f'B{row}'].fill = inv_fill
        for col in ['A', 'B']:
            ws[f'{col}{row}'].border = thin_border
        row += 2

        # Financing Activities
        ws[f'A{row}'] = "FINANCING ACTIVITIES"
        ws[f'A{row}'].font = section_font
        row += 1

        ws[f'A{row}'] = "Item"
        ws[f'B{row}'] = "Amount"
        for col in ['A', 'B']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].border = thin_border
        row += 1

        financing = data.get('financing_activities', {})
        fin_items = [
            ('Loan Proceeds', financing.get('loan_proceeds', 0)),
            ('Loan Repayments', -financing.get('loan_repayments', 0)),
            ('Equity Injections', financing.get('equity_injections', 0)),
            ('Dividends Paid', -financing.get('dividends_paid', 0)),
            ('Other Financing Cash', financing.get('other', 0)),
        ]

        for name, value in fin_items:
            ws[f'A{row}'] = name
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = currency_format
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = thin_border
            row += 1

        ws[f'A{row}'] = "NET CASH FROM FINANCING"
        ws[f'B{row}'] = financing.get('net_cash_from_financing', 0)
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].font = total_font
        ws[f'B{row}'].font = total_font
        fin_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
        ws[f'A{row}'].fill = fin_fill
        ws[f'B{row}'].fill = fin_fill
        for col in ['A', 'B']:
            ws[f'{col}{row}'].border = thin_border
        row += 2

        # Summary
        ws[f'A{row}'] = "SUMMARY"
        ws[f'A{row}'].font = section_font
        row += 1

        ws[f'A{row}'] = "Net Change in Cash"
        ws[f'B{row}'] = data.get('net_change_in_cash', 0)
        ws[f'B{row}'].number_format = currency_format
        row += 1

        ws[f'A{row}'] = "Opening Cash Balance"
        ws[f'B{row}'] = data.get('opening_cash_balance', 0)
        ws[f'B{row}'].number_format = currency_format
        row += 1

        ws[f'A{row}'] = "CLOSING CASH BALANCE"
        ws[f'B{row}'] = data.get('closing_cash_balance', 0)
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'].font = Font(bold=True, size=12)

        closing = data.get('closing_cash_balance', 0)
        closing_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid") if closing >= 0 else PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        ws[f'A{row}'].fill = closing_fill
        ws[f'B{row}'].fill = closing_fill

        # Auto-fit columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # =========================================================================
    # BALANCE SHEET EXPORTS
    # =========================================================================

    def generate_balance_sheet_pdf(
        self,
        data: Dict,
        year: int,
        quarter: int,
        billing_entity: str = 'all',
        currency: str = 'all'
    ) -> BytesIO:
        """
        Generate PDF for Balance Sheet.

        Args:
            data: Balance Sheet data from BalanceSheetService
            year: Year
            quarter: Quarter (1-4)
            billing_entity: Billing entity filter
            currency: Currency filter

        Returns:
            BytesIO buffer containing the PDF
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=letter,
            topMargin=0.5 * inch, bottomMargin=0.5 * inch,
            leftMargin=0.5 * inch, rightMargin=0.5 * inch
        )

        elements = []

        # Period text
        quarter_name = data.get('period', {}).get('quarter_name', f'Q{quarter}')
        as_of_date = data.get('period', {}).get('as_of_date', '')
        period_text = f"{quarter_name} {year} (As of {as_of_date})"

        # Header
        self._add_pdf_header(
            elements, "BALANCE SHEET",
            period_text, billing_entity, currency
        )

        curr = currency if currency != 'all' else 'THB'

        # ASSETS Section
        elements.append(Paragraph("ASSETS", self.heading_style))

        assets = data.get('assets', {})

        # Current Assets
        elements.append(Paragraph("Current Assets", self.section_style))
        current = assets.get('current_assets', {})
        current_data = [
            ['Item', 'Amount'],
            ['Cash & Bank', self._format_currency(current.get('cash_and_bank', 0), curr)],
            ['Accounts Receivable', self._format_currency(current.get('accounts_receivable', {}).get('total', 0), curr)],
            ['Other Current Assets', self._format_currency(current.get('other_current_assets', 0), curr)],
            ['TOTAL CURRENT ASSETS', self._format_currency(current.get('total', 0), curr)],
        ]

        current_table = Table(current_data, colWidths=[5 * inch, 2 * inch])
        current_table.setStyle(self._create_table_style())
        current_table.setStyle(TableStyle([
            ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 10),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e3f2fd')),
        ]))
        elements.append(current_table)
        elements.append(Spacer(1, 0.1 * inch))

        # Fixed Assets
        elements.append(Paragraph("Fixed Assets", self.section_style))
        fixed = assets.get('fixed_assets', {})
        fixed_data = [
            ['Item', 'Amount'],
            ['Gross Fixed Assets', self._format_currency(fixed.get('gross_fixed_assets', {}).get('total', 0), curr)],
            ['Less: Accumulated Depreciation', f"({self._format_currency(fixed.get('accumulated_depreciation', 0), curr)})"],
            ['NET FIXED ASSETS', self._format_currency(fixed.get('net_fixed_assets', 0), curr)],
        ]

        fixed_table = Table(fixed_data, colWidths=[5 * inch, 2 * inch])
        fixed_table.setStyle(self._create_table_style())
        fixed_table.setStyle(TableStyle([
            ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 10),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f5e9')),
        ]))
        elements.append(fixed_table)
        elements.append(Spacer(1, 0.1 * inch))

        # Total Assets
        elements.append(Paragraph(
            f"<b>TOTAL ASSETS:</b> {self._format_currency(assets.get('total_assets', 0), curr)}",
            self.body_style
        ))
        elements.append(Spacer(1, 0.2 * inch))

        # LIABILITIES Section
        elements.append(Paragraph("LIABILITIES", self.heading_style))

        liabilities = data.get('liabilities', {})

        # Current Liabilities
        elements.append(Paragraph("Current Liabilities", self.section_style))
        curr_liab = liabilities.get('current_liabilities', {})
        curr_liab_data = [
            ['Item', 'Amount'],
            ['Accounts Payable', self._format_currency(curr_liab.get('accounts_payable', {}).get('total', 0), curr)],
            ['Accrued Expenses', self._format_currency(curr_liab.get('accrued_expenses', 0), curr)],
            ['Other Current Liabilities', self._format_currency(curr_liab.get('other_current_liabilities', 0), curr)],
            ['TOTAL CURRENT LIABILITIES', self._format_currency(curr_liab.get('total', 0), curr)],
        ]

        curr_liab_table = Table(curr_liab_data, colWidths=[5 * inch, 2 * inch])
        curr_liab_table.setStyle(self._create_table_style())
        curr_liab_table.setStyle(TableStyle([
            ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 10),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ffebee')),
        ]))
        elements.append(curr_liab_table)
        elements.append(Spacer(1, 0.1 * inch))

        # Long-term Liabilities
        elements.append(Paragraph("Long-term Liabilities", self.section_style))
        lt_liab = liabilities.get('long_term_liabilities', {})
        lt_liab_data = [
            ['Item', 'Amount'],
            ['Long-term Debt', self._format_currency(lt_liab.get('long_term_debt', 0), curr)],
            ['Other Long-term Liabilities', self._format_currency(lt_liab.get('other_long_term_liabilities', 0), curr)],
            ['TOTAL LONG-TERM LIABILITIES', self._format_currency(lt_liab.get('total', 0), curr)],
        ]

        lt_liab_table = Table(lt_liab_data, colWidths=[5 * inch, 2 * inch])
        lt_liab_table.setStyle(self._create_table_style())
        lt_liab_table.setStyle(TableStyle([
            ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 10),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fce4ec')),
        ]))
        elements.append(lt_liab_table)
        elements.append(Spacer(1, 0.1 * inch))

        # Total Liabilities
        elements.append(Paragraph(
            f"<b>TOTAL LIABILITIES:</b> {self._format_currency(liabilities.get('total_liabilities', 0), curr)}",
            self.body_style
        ))
        elements.append(Spacer(1, 0.2 * inch))

        # EQUITY Section
        elements.append(Paragraph("EQUITY", self.heading_style))

        equity = data.get('equity', {})
        equity_data = [
            ['Item', 'Amount'],
            ['Share Capital', self._format_currency(equity.get('share_capital', 0), curr)],
            ['Additional Paid-in Capital', self._format_currency(equity.get('additional_paid_in_capital', 0), curr)],
            ['Retained Earnings', self._format_currency(equity.get('retained_earnings', 0), curr)],
            ['Other Equity', self._format_currency(equity.get('other_equity', 0), curr)],
            ['TOTAL EQUITY', self._format_currency(equity.get('total_equity', 0), curr)],
        ]

        equity_table = Table(equity_data, colWidths=[5 * inch, 2 * inch])
        equity_table.setStyle(self._create_table_style())
        equity_table.setStyle(TableStyle([
            ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 10),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f3e5f5')),
        ]))
        elements.append(equity_table)
        elements.append(Spacer(1, 0.2 * inch))

        # Balance Check
        elements.append(HRFlowable(
            width="100%", thickness=1,
            color=colors.HexColor('#FFA500'),
            spaceBefore=0, spaceAfter=0
        ))
        elements.append(Spacer(1, 0.15 * inch))

        total_assets = assets.get('total_assets', 0)
        total_l_and_e = data.get('total_liabilities_and_equity', 0)
        is_balanced = data.get('is_balanced', False)

        balance_data = [
            ['Description', 'Amount'],
            ['Total Assets', self._format_currency(total_assets, curr)],
            ['Total Liabilities + Equity', self._format_currency(total_l_and_e, curr)],
            ['Balance Status', 'BALANCED' if is_balanced else f"UNBALANCED (Diff: {self._format_currency(data.get('balance_difference', 0), curr)})"],
        ]

        balance_table = Table(balance_data, colWidths=[5 * inch, 2 * inch])
        balance_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFA500')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONT', (0, 1), (-1, -1), 'Helvetica', 10),
            ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 10),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f5e9') if is_balanced else colors.HexColor('#ffebee')),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e0')),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(balance_table)

        # Footer
        self._add_pdf_footer(elements)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer

    def generate_balance_sheet_excel(
        self,
        data: Dict,
        year: int,
        quarter: int,
        billing_entity: str = 'all',
        currency: str = 'all'
    ) -> BytesIO:
        """
        Generate Excel for Balance Sheet.

        Args:
            data: Balance Sheet data from BalanceSheetService
            year: Year
            quarter: Quarter (1-4)
            billing_entity: Billing entity filter
            currency: Currency filter

        Returns:
            BytesIO buffer containing the Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"

        # Styles
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        section_font = Font(bold=True, size=11)
        total_font = Font(bold=True, size=10)
        currency_format = '#,##0.00'
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        curr = currency if currency != 'all' else 'THB'
        entity_info = self._get_entity_info(billing_entity)

        # Period text
        quarter_name = data.get('period', {}).get('quarter_name', f'Q{quarter}')
        as_of_date = data.get('period', {}).get('as_of_date', '')
        period_text = f"{quarter_name} {year} (As of {as_of_date})"

        # Header section
        ws['A1'] = "BALANCE SHEET"
        ws['A1'].font = Font(bold=True, size=16, color="FFA500")
        ws.merge_cells('A1:B1')

        ws['A2'] = f"Period: {period_text}"
        ws['A3'] = f"Entity: {entity_info['name']}"
        ws['A4'] = f"Currency: {curr}"
        ws['A5'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"

        row = 7

        # ASSETS
        ws[f'A{row}'] = "ASSETS"
        ws[f'A{row}'].font = section_font
        row += 1

        # Current Assets
        ws[f'A{row}'] = "Current Assets"
        ws[f'A{row}'].font = Font(italic=True, bold=True)
        row += 1

        ws[f'A{row}'] = "Item"
        ws[f'B{row}'] = "Amount"
        for col in ['A', 'B']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].border = thin_border
        row += 1

        assets = data.get('assets', {})
        current = assets.get('current_assets', {})

        current_items = [
            ('Cash & Bank', current.get('cash_and_bank', 0)),
            ('Accounts Receivable', current.get('accounts_receivable', {}).get('total', 0)),
            ('Other Current Assets', current.get('other_current_assets', 0)),
        ]

        for name, value in current_items:
            ws[f'A{row}'] = name
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = currency_format
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = thin_border
            row += 1

        ws[f'A{row}'] = "TOTAL CURRENT ASSETS"
        ws[f'B{row}'] = current.get('total', 0)
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].font = total_font
        ws[f'B{row}'].font = total_font
        curr_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        ws[f'A{row}'].fill = curr_fill
        ws[f'B{row}'].fill = curr_fill
        for col in ['A', 'B']:
            ws[f'{col}{row}'].border = thin_border
        row += 2

        # Fixed Assets
        ws[f'A{row}'] = "Fixed Assets"
        ws[f'A{row}'].font = Font(italic=True, bold=True)
        row += 1

        fixed = assets.get('fixed_assets', {})
        fixed_items = [
            ('Gross Fixed Assets', fixed.get('gross_fixed_assets', {}).get('total', 0)),
            ('Less: Accumulated Depreciation', -fixed.get('accumulated_depreciation', 0)),
        ]

        for name, value in fixed_items:
            ws[f'A{row}'] = name
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = currency_format
            row += 1

        ws[f'A{row}'] = "NET FIXED ASSETS"
        ws[f'B{row}'] = fixed.get('net_fixed_assets', 0)
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].font = total_font
        ws[f'B{row}'].font = total_font
        fixed_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        ws[f'A{row}'].fill = fixed_fill
        ws[f'B{row}'].fill = fixed_fill
        row += 2

        ws[f'A{row}'] = "TOTAL ASSETS"
        ws[f'B{row}'] = assets.get('total_assets', 0)
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'].font = Font(bold=True, size=12)
        row += 2

        # LIABILITIES
        ws[f'A{row}'] = "LIABILITIES"
        ws[f'A{row}'].font = section_font
        row += 1

        liabilities = data.get('liabilities', {})

        # Current Liabilities
        ws[f'A{row}'] = "Current Liabilities"
        ws[f'A{row}'].font = Font(italic=True, bold=True)
        row += 1

        curr_liab = liabilities.get('current_liabilities', {})
        cl_items = [
            ('Accounts Payable', curr_liab.get('accounts_payable', {}).get('total', 0)),
            ('Accrued Expenses', curr_liab.get('accrued_expenses', 0)),
            ('Other Current Liabilities', curr_liab.get('other_current_liabilities', 0)),
        ]

        for name, value in cl_items:
            ws[f'A{row}'] = name
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = currency_format
            row += 1

        ws[f'A{row}'] = "TOTAL CURRENT LIABILITIES"
        ws[f'B{row}'] = curr_liab.get('total', 0)
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].font = total_font
        ws[f'B{row}'].font = total_font
        cl_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        ws[f'A{row}'].fill = cl_fill
        ws[f'B{row}'].fill = cl_fill
        row += 2

        # Long-term Liabilities
        ws[f'A{row}'] = "Long-term Liabilities"
        ws[f'A{row}'].font = Font(italic=True, bold=True)
        row += 1

        lt_liab = liabilities.get('long_term_liabilities', {})
        lt_items = [
            ('Long-term Debt', lt_liab.get('long_term_debt', 0)),
            ('Other Long-term Liabilities', lt_liab.get('other_long_term_liabilities', 0)),
        ]

        for name, value in lt_items:
            ws[f'A{row}'] = name
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = currency_format
            row += 1

        ws[f'A{row}'] = "TOTAL LONG-TERM LIABILITIES"
        ws[f'B{row}'] = lt_liab.get('total', 0)
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].font = total_font
        ws[f'B{row}'].font = total_font
        row += 2

        ws[f'A{row}'] = "TOTAL LIABILITIES"
        ws[f'B{row}'] = liabilities.get('total_liabilities', 0)
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'].font = Font(bold=True, size=12)
        row += 2

        # EQUITY
        ws[f'A{row}'] = "EQUITY"
        ws[f'A{row}'].font = section_font
        row += 1

        equity = data.get('equity', {})
        equity_items = [
            ('Share Capital', equity.get('share_capital', 0)),
            ('Additional Paid-in Capital', equity.get('additional_paid_in_capital', 0)),
            ('Retained Earnings', equity.get('retained_earnings', 0)),
            ('Other Equity', equity.get('other_equity', 0)),
        ]

        for name, value in equity_items:
            ws[f'A{row}'] = name
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = currency_format
            row += 1

        ws[f'A{row}'] = "TOTAL EQUITY"
        ws[f'B{row}'] = equity.get('total_equity', 0)
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].font = total_font
        ws[f'B{row}'].font = total_font
        eq_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
        ws[f'A{row}'].fill = eq_fill
        ws[f'B{row}'].fill = eq_fill
        row += 2

        # Balance Check
        ws[f'A{row}'] = "BALANCE CHECK"
        ws[f'A{row}'].font = section_font
        row += 1

        ws[f'A{row}'] = "Total Assets"
        ws[f'B{row}'] = assets.get('total_assets', 0)
        ws[f'B{row}'].number_format = currency_format
        row += 1

        ws[f'A{row}'] = "Total Liabilities + Equity"
        ws[f'B{row}'] = data.get('total_liabilities_and_equity', 0)
        ws[f'B{row}'].number_format = currency_format
        row += 1

        is_balanced = data.get('is_balanced', False)
        ws[f'A{row}'] = "Balance Status"
        ws[f'B{row}'] = "BALANCED" if is_balanced else f"UNBALANCED (Diff: {data.get('balance_difference', 0):.2f})"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'].font = Font(bold=True, size=12)
        balance_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid") if is_balanced else PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        ws[f'A{row}'].fill = balance_fill
        ws[f'B{row}'].fill = balance_fill

        # Auto-fit columns
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
